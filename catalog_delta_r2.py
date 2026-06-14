#!/usr/bin/env python3
"""
catalog_delta_r2.py
Erzeugt catalog_review_delta_r2.xlsx mit ausschließlich den neuen Einträgen
aus Round 2 (catalog_raw_r2/) und catalog_manual.json.
Gleiche Spalten wie catalog_review.xlsx — kann unabhängig reviewt werden.
"""
import json, pathlib, sys, openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

REPO_ROOT       = pathlib.Path(__file__).parent
CATALOG_RAW_R2  = REPO_ROOT / "catalog_raw_r2"
CATALOG_FULL    = REPO_ROOT / "catalog_full.json"
CATALOG_MANUAL  = REPO_ROOT / "catalog_manual.json"
OUT_DELTA       = REPO_ROOT / "catalog_review_delta_r2.xlsx"

XLSX_COLS = [
    "production_rank", "thema", "themengebiet", "tier", "leuchtturm",
    "erg_s1", "erg_s2", "erg_s3",
    "eignung", "age_floor", "kategorie_nr", "framing_note",
    "sensibel", "begruendung_eignung", "dublette_von", "notiz",
    "FREIGABE",
]
FILL_HEADER   = PatternFill("solid", fgColor="1F4E79")
FILL_SENSIBEL = PatternFill("solid", fgColor="FFD7D7")
FILL_MANUAL   = PatternFill("solid", fgColor="E2EFDA")
FONT_HEADER   = Font(color="FFFFFF", bold=True)

def main():
    r2_themen: set[str] = set()
    manual_themen: set[str] = set()

    if CATALOG_RAW_R2.exists():
        for f in sorted(CATALOG_RAW_R2.glob("*.json")):
            if "_raw" in f.name:
                continue
            for item in json.loads(f.read_text(encoding="utf-8")):
                t = item.get("thema", "").strip()
                if t:
                    r2_themen.add(t)

    if CATALOG_MANUAL.exists():
        for item in json.loads(CATALOG_MANUAL.read_text(encoding="utf-8")):
            t = item.get("thema", "").strip()
            if t:
                r2_themen.add(t)
                manual_themen.add(t)

    print(f"R2-Quell-Themen (roh): {len(r2_themen)}")

    full = json.loads(CATALOG_FULL.read_text(encoding="utf-8"))
    delta = [x for x in full if x.get("thema", "").strip() in r2_themen]
    delta.sort(key=lambda x: x.get("production_rank", 9999))
    print(f"Davon primary in catalog_full.json: {len(delta)}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Review_R2"

    for ci, col in enumerate(XLSX_COLS, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = Alignment(horizontal="center")

    sorted_items = sorted(delta, key=lambda x: (
        0 if x.get("sensibel") else 1,
        x.get("production_rank", 9999),
    ))
    for ri, item in enumerate(sorted_items, 2):
        for ci, col in enumerate(XLSX_COLS, 1):
            if col == "FREIGABE":
                ws.cell(row=ri, column=ci, value="")
            else:
                v = item.get(col)
                ws.cell(row=ri, column=ci, value="TRUE" if v is True else ("" if v is False else v))

        thema = item.get("thema", "")
        if item.get("sensibel"):
            fill = FILL_SENSIBEL
        elif thema in manual_themen:
            fill = FILL_MANUAL
        else:
            fill = None
        if fill:
            for ci in range(1, len(XLSX_COLS) + 1):
                ws.cell(row=ri, column=ci).fill = fill

    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["L"].width = 40
    ws.column_dimensions["N"].width = 45
    ws.column_dimensions["Q"].width = 15
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:Q{len(sorted_items)+1}"

    wb.save(OUT_DELTA)
    n_sensibel = sum(1 for x in delta if x.get("sensibel"))
    n_manual   = sum(1 for x in delta if x.get("thema", "") in manual_themen)
    print(f"→ {OUT_DELTA.name}: {len(delta)} Einträge ({n_sensibel} sensibel, {n_manual} manuell/grün)")

if __name__ == "__main__":
    main()
