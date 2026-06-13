#!/usr/bin/env python3
"""
catalog_merge.py  v1  (2026-06-13)
───────────────────────────────────────────────────────────────────────────
Wissensfreund — Katalog-Merge

Lädt alle catalog_raw/*.json, dedupliziert über Gebietsgrenzen hinweg,
weist Produktions-Reihenfolge zu und exportiert:
  - catalog_full.json      (alle include-Themen, gerankt)
  - catalog_reserve.json   (reserve-Themen separat)
  - catalog_review.xlsx    (für menschliche Freigabe; sensibel=true oben/rot)

Voraussetzungen:
  pip install openpyxl
"""

import json
import pathlib
import sys
from collections import defaultdict

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# ── Pfade ─────────────────────────────────────────────────────────────────

REPO_ROOT       = pathlib.Path(__file__).parent
CATALOG_RAW_DIR = REPO_ROOT / "catalog_raw"
OUT_JSON        = REPO_ROOT / "catalog_full.json"
OUT_RESERVE     = REPO_ROOT / "catalog_reserve.json"
OUT_XLSX        = REPO_ROOT / "catalog_review.xlsx"

# Spalten-Reihenfolge im Excel
XLSX_COLS = [
    "production_rank", "thema", "themengebiet", "tier", "leuchtturm",
    "erg_s1", "erg_s2", "erg_s3",
    "eignung", "age_floor", "kategorie_nr", "framing_note",
    "sensibel", "begruendung_eignung", "dublette_von", "notiz",
    "FREIGABE",   # ← manuell: leer lassen = Rater-Urteil übernehmen; "OK" / "EXCLUDE" / "AGE_FLOOR=2" usw.
]


# ── 1. Laden ───────────────────────────────────────────────────────────────

def load_all() -> list[dict]:
    items = []
    files = sorted(f for f in CATALOG_RAW_DIR.glob("*.json") if "_raw" not in f.name)
    for f in files:
        data = json.loads(f.read_text(encoding="utf-8"))
        items.extend(data)
    print(f"  {len(items)} Items aus {len(files)} Datei(en) geladen.")
    return items


# ── 2. Dedup ───────────────────────────────────────────────────────────────

def normalize(s: str) -> str:
    return s.strip().lower()


def priority(item: dict) -> tuple:
    """Höhere Priorität = wird bei Duplikaten behalten."""
    return (
        1 if item.get("tier") == "primary" else 0,
        item.get("erg_s2") or 0,
        item.get("erg_s3") or 0,
        item.get("erg_s1") or 0,
    )


def dedup(items: list[dict]) -> tuple[list[dict], list[dict]]:
    """
    Dedup nach normalisiertem thema-Lemma.
    Berücksichtigt zusätzlich das vom Rater gesetzte dublette_von-Feld.
    Gibt (kanonische Items, Dubletten-Liste) zurück.
    """
    # Rater-markierte Dubletten zuerst herausfiltern
    rater_dupes = {normalize(x["thema"]) for x in items if x.get("dublette_von")}

    seen: dict[str, dict] = {}
    duplicates: list[dict] = []

    for item in items:
        key = normalize(item.get("thema", ""))
        if not key:
            continue

        # Vom Rater als Dublette markiert → direkt in Duplikat-Liste
        if key in rater_dupes and item.get("dublette_von"):
            duplicates.append(item)
            continue

        if key not in seen:
            seen[key] = item
        else:
            existing = seen[key]
            if priority(item) > priority(existing):
                duplicates.append({**existing, "_merge_duplikat": True})
                seen[key] = item
            else:
                duplicates.append({**item, "_merge_duplikat": True})

    return list(seen.values()), duplicates


# ── 3. Produktions-Reihenfolge ─────────────────────────────────────────────

def assign_ranks(items: list[dict]) -> tuple[list[dict], list[dict]]:
    """
    Trennt primary von reserve, weist production_rank zu.

    Primary-Reihenfolge:
      1. Leuchtturm-Themen (erg_s2 desc)
      2. Übrige primary-Themen — Round-Robin über Themengebiete (erg_s2 desc je Gebiet)

    Reserve: separat, nach erg_s2 desc — kein production_rank.
    Exclude-Themen: keine production_rank, nicht in catalog_full.json.
    """
    primary = [x for x in items if x.get("tier") == "primary" and x.get("eignung") != "exclude"]
    reserve = [x for x in items if x.get("tier") == "reserve"]
    # exclude-Themen bleiben im XLSX sichtbar, kommen nicht in catalog_full.json

    # Leuchtturm-Block
    leuchtturm = sorted(
        [x for x in primary if x.get("leuchtturm")],
        key=lambda x: (x.get("erg_s2") or 0, x.get("erg_s3") or 0),
        reverse=True,
    )

    # Round-Robin nicht-Leuchtturm
    non_leu = [x for x in primary if not x.get("leuchtturm")]
    by_gebiet: dict[str, list[dict]] = defaultdict(list)
    for item in non_leu:
        by_gebiet[item.get("themengebiet", "—")].append(item)
    for g in by_gebiet:
        by_gebiet[g].sort(key=lambda x: (x.get("erg_s2") or 0, x.get("erg_s3") or 0), reverse=True)

    gebiete = sorted(by_gebiet.keys())
    ptrs = {g: 0 for g in gebiete}
    rr: list[dict] = []
    while any(ptrs[g] < len(by_gebiet[g]) for g in gebiete):
        for g in gebiete:
            if ptrs[g] < len(by_gebiet[g]):
                rr.append(by_gebiet[g][ptrs[g]])
                ptrs[g] += 1

    ordered = leuchtturm + rr
    for i, item in enumerate(ordered, 1):
        item["production_rank"] = i

    # Reserve: sortiert, kein Rank
    reserve.sort(key=lambda x: (x.get("erg_s2") or 0, x.get("erg_s3") or 0), reverse=True)

    return ordered, reserve


# ── 4. JSON-Export ─────────────────────────────────────────────────────────

def export_json(primary: list[dict], reserve: list[dict]) -> None:
    OUT_JSON.write_text(json.dumps(primary, ensure_ascii=False, indent=2), encoding="utf-8")
    OUT_RESERVE.write_text(json.dumps(reserve, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"  catalog_full.json:    {len(primary)} Themen (Rank 1–{len(primary)})")
    print(f"  catalog_reserve.json: {len(reserve)} Reserve-Themen")


# ── 5. Excel-Export ────────────────────────────────────────────────────────

FILL_HEADER  = PatternFill("solid", fgColor="1F4E79")
FILL_SENSIBEL = PatternFill("solid", fgColor="FFD7D7")
FILL_EXCLUDE  = PatternFill("solid", fgColor="FF9999")
FILL_RESERVE  = PatternFill("solid", fgColor="F2F2F2")
FILL_LEUCHT   = PatternFill("solid", fgColor="FFF2CC")
FONT_HEADER   = Font(color="FFFFFF", bold=True)
FONT_BOLD     = Font(bold=True)


def cell_value(item: dict, col: str):
    if col == "FREIGABE":
        return ""
    v = item.get(col)
    if isinstance(v, bool):
        return "TRUE" if v else ""
    return v


def export_xlsx(primary: list[dict], reserve: list[dict], duplicates: list[dict], all_items: list[dict]) -> None:
    wb = openpyxl.Workbook()

    # ── Sheet 1: Review ───────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Review"

    # Header
    for ci, col in enumerate(XLSX_COLS, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = Alignment(horizontal="center")

    # Sortierung: eignung=exclude zuerst, dann sensibel=true, dann production_rank
    def sort_key(x):
        return (
            0 if x.get("eignung") == "exclude" else 1,
            0 if x.get("sensibel") else 1,
            x.get("production_rank") or 99999,
        )

    all_for_review = sorted(all_items, key=sort_key)

    for ri, item in enumerate(all_for_review, 2):
        for ci, col in enumerate(XLSX_COLS, 1):
            c = ws.cell(row=ri, column=ci, value=cell_value(item, col))

        # Zeilenfarbe
        if item.get("eignung") == "exclude":
            fill = FILL_EXCLUDE
        elif item.get("sensibel"):
            fill = FILL_SENSIBEL
        elif item.get("tier") == "reserve":
            fill = FILL_RESERVE
        elif item.get("leuchtturm"):
            fill = FILL_LEUCHT
        else:
            fill = None

        if fill:
            for ci in range(1, len(XLSX_COLS) + 1):
                ws.cell(row=ri, column=ci).fill = fill

    # Spaltenbreiten
    col_widths = {
        "A": 8,   # rank
        "B": 28,  # thema
        "C": 22,  # themengebiet
        "D": 9,   # tier
        "E": 10,  # leuchtturm
        "F": 5, "G": 5, "H": 5,  # erg
        "I": 10,  # eignung
        "J": 8,   # age_floor
        "K": 12,  # kategorie_nr
        "L": 42,  # framing_note
        "M": 10,  # sensibel
        "N": 45,  # begruendung_eignung
        "O": 20,  # dublette_von
        "P": 25,  # notiz
        "Q": 15,  # FREIGABE
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:Q{len(all_for_review) + 1}"

    # ── Sheet 2: Statistik ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Statistik")
    ws2.append(["Themengebiet", "Total", "Primary", "Reserve", "Sensibel", "Leuchtturm", "Exclude"])
    ws2["A1"].font = FONT_BOLD

    by_g: dict[str, list[dict]] = defaultdict(list)
    for item in all_items:
        by_g[item.get("themengebiet", "—")].append(item)

    totals = [0, 0, 0, 0, 0, 0]
    for g in sorted(by_g.keys()):
        g_items = by_g[g]
        row = [
            g,
            len(g_items),
            sum(1 for x in g_items if x.get("tier") == "primary"),
            sum(1 for x in g_items if x.get("tier") == "reserve"),
            sum(1 for x in g_items if x.get("sensibel")),
            sum(1 for x in g_items if x.get("leuchtturm")),
            sum(1 for x in g_items if x.get("eignung") == "exclude"),
        ]
        ws2.append(row)
        for i, v in enumerate(row[1:], 0):
            totals[i] += v

    ws2.append(["GESAMT"] + totals)
    last_row = ws2.max_row
    for ci in range(1, 8):
        ws2.cell(row=last_row, column=ci).font = FONT_BOLD

    ws2.column_dimensions["A"].width = 30

    # ── Sheet 3: Dubletten ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Dubletten")
    dup_cols = ["thema", "themengebiet", "tier", "erg_s2", "dublette_von", "_merge_duplikat"]
    for ci, col in enumerate(dup_cols, 1):
        ws3.cell(row=1, column=ci, value=col).font = FONT_BOLD
    for ri, dup in enumerate(duplicates, 2):
        for ci, col in enumerate(dup_cols, 1):
            ws3.cell(row=ri, column=ci, value=dup.get(col))

    if not duplicates:
        ws3.cell(row=2, column=1, value="Keine Dubletten gefunden.")

    wb.save(OUT_XLSX)
    print(f"  catalog_review.xlsx:  {len(all_for_review)} Zeilen"
          f" ({sum(1 for x in all_for_review if x.get('sensibel'))} sensibel,"
          f" {sum(1 for x in all_for_review if x.get('eignung')=='exclude')} exclude)")


# ── Main ───────────────────────────────────────────────────────────────────

def main() -> None:
    print("1. Lade …")
    all_items = load_all()

    print("2. Dedup …")
    canonical, duplicates = dedup(all_items)
    print(f"  {len(duplicates)} Dubletten → {len(canonical)} eindeutige Themen")

    print("3. Produktions-Reihenfolge …")
    primary, reserve = assign_ranks(canonical)
    n_leu  = sum(1 for x in primary if x.get("leuchtturm"))
    n_excl = sum(1 for x in canonical if x.get("eignung") == "exclude")
    n_sens = sum(1 for x in canonical if x.get("sensibel"))
    print(f"  {len(primary)} primary (davon {n_leu} Leuchtturm, erste 500 = Rank 1–500)")
    print(f"  {len(reserve)} reserve")
    print(f"  {n_excl} exclude (im XLSX markiert, nicht in catalog_full.json)")
    print(f"  {n_sens} sensibel → Freigabe erforderlich")

    print("4. Exportiere …")
    export_json(primary, reserve)
    export_xlsx(primary, reserve, duplicates, canonical)

    print("\n✓ Fertig.")
    print(f"  Nächster Schritt: catalog_review.xlsx öffnen, sensibel/exclude-Zeilen prüfen,")
    print(f"  FREIGABE-Spalte ausfüllen → eignung_verdicts.json generieren.")


if __name__ == "__main__":
    main()
