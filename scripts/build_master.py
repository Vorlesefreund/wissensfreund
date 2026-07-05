#!/usr/bin/env python3
"""
build_master.py  —  Erzeugt catalog_review_master.xlsx frisch aus aktuellem Katalog-Stand.

Quellen:
  - catalog_merge.py: load_all(), dedup(), apply_master_annotations()
  - catalog_review_master.xlsx: bestehende Annotierungen (eignung, erg, age_floor, etc.)

Sortierung: themengebiet alpha, dann thema alpha — für menschliches Review.
Orange thema-Zelle: erg unvollständig (erg_s1 oder erg_s2 fehlt, oder erg_s3 fehlt
  wenn age_floor >= 3 oder generell).
Neue Spalte: Kommentar (freies Feld für Andreas).

Aufruf: python -X utf8 scripts/build_master.py
"""

import pathlib, sys, json
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

# catalog_merge aus Repo-Root importieren
REPO = pathlib.Path(__file__).parent.parent
sys.path.insert(0, str(REPO))
import catalog_merge as cm

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

# ── Farben ────────────────────────────────────────────────────────────────────
FILL_HEADER   = PatternFill("solid", fgColor="1F4E79")
FILL_EXCLUDE  = PatternFill("solid", fgColor="FF9999")
FILL_SENSIBEL = PatternFill("solid", fgColor="FFD7D7")
FILL_LEUCHT   = PatternFill("solid", fgColor="FFF2CC")
FILL_RESERVE  = PatternFill("solid", fgColor="F2F2F2")
FILL_ERG_WARN = PatternFill("solid", fgColor="FFD966")   # orange: unvollständige erg
FONT_HEADER   = Font(color="FFFFFF", bold=True)
FONT_BOLD     = Font(bold=True)

COLS = [
    "production_rank", "thema", "themengebiet", "tier", "leuchtturm",
    "erg_s1", "erg_s2", "erg_s3",
    "eignung", "age_floor", "kategorie_nr", "framing_note",
    "sensibel", "begruendung_eignung", "dublette_von", "notiz",
    "FREIGABE", "Kommentar", "themengebiete",
]

COL_WIDTHS = {
    "A": 8,   # rank
    "B": 28,  # thema
    "C": 22,  # themengebiet
    "D": 9,   # tier
    "E": 10,  # leuchtturm
    "F": 5, "G": 5, "H": 5,    # erg
    "I": 10,  # eignung
    "J": 8,   # age_floor
    "K": 12,  # kategorie_nr
    "L": 42,  # framing_note
    "M": 10,  # sensibel
    "N": 45,  # begruendung_eignung
    "O": 20,  # dublette_von
    "P": 25,  # notiz
    "Q": 15,  # FREIGABE
    "R": 35,  # Kommentar
    "S": 34,  # themengebiete
}


def erg_incomplete(item: dict) -> bool:
    """True wenn ein ab age_floor benötigtes erg-Feld fehlt."""
    if item.get("eignung") == "exclude" or item.get("tier") == "reserve":
        return False
    try:
        af = int(item.get("age_floor") or 1)
    except Exception:
        af = 1
    for i, s in enumerate(["erg_s1", "erg_s2", "erg_s3"], start=1):
        if i >= af and item.get(s) in (None, "", 0):
            return True
    return False


def main() -> None:
    OUT = REPO / "catalog_review_master.xlsx"

    # 1. Alle Items laden + dedup
    print("1. Lade + dedup …")
    all_items = cm.load_all()
    canonical, _ = cm.dedup(all_items)
    print(f"   {len(canonical)} eindeutige Themen")

    # 2. Master-Annotierungen anwenden
    print("2. Master-Annotierungen laden …")
    master_ann = cm.load_existing_annotations(OUT)
    n = cm.apply_master_annotations(canonical, master_ann)
    print(f"   {n} Themen mit Override")
    n_tg = cm.apply_themengebiete_annotations(canonical)
    print(f"   {n_tg} Themen mit themengebiete-Liste (inkl. ggf. Primär-Umschichtung)")

    # 3. production_rank zuweisen (für Referenz, aber Sortierung im xlsx ist alpha)
    primary, reserve = cm.assign_ranks(canonical)
    rank_map = {x["thema"]: x.get("production_rank") for x in primary}

    # 4. Sortierung: themengebiet alpha, dann thema alpha
    def sort_key(x):
        return (x.get("themengebiet") or "zzz", x.get("thema") or "")

    all_sorted = sorted(canonical, key=sort_key)

    n_excl  = sum(1 for x in canonical if x.get("eignung") == "exclude")
    n_sens  = sum(1 for x in canonical if x.get("sensibel"))
    n_leu   = sum(1 for x in canonical if x.get("leuchtturm"))
    n_orange = sum(1 for x in canonical if erg_incomplete(x) and x.get("eignung") != "exclude")
    print(f"   {len(all_sorted)} Zeilen gesamt | {n_excl} exclude | {n_sens} sensibel | {n_leu} Leuchtturm | {n_orange} erg-orange")

    # 5. Excel schreiben
    print("3. Schreibe catalog_review_master.xlsx …")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Review"

    # Header
    for ci, col in enumerate(COLS, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = Alignment(horizontal="center")

    # Daten
    for ri, item in enumerate(all_sorted, 2):
        thema = item.get("thema", "")
        ann = master_ann.get(thema.strip(), {})
        is_excl = item.get("eignung") == "exclude"
        is_sens = bool(item.get("sensibel"))
        is_leu  = bool(item.get("leuchtturm"))
        is_res  = item.get("tier") == "reserve"
        erg_warn = erg_incomplete(item) and not is_excl

        for ci, col in enumerate(COLS, 1):
            if col == "production_rank":
                v = rank_map.get(thema)
            elif col == "FREIGABE":
                v = ann.get("FREIGABE", "")
            elif col == "Kommentar":
                v = ann.get("Kommentar", "")
            elif col == "themengebiete":
                tg = item.get("themengebiete") or []
                v = "|".join(tg) if tg else ""
            else:
                v = item.get(col)
                if isinstance(v, bool):
                    v = "TRUE" if v else ""
            ws.cell(row=ri, column=ci, value=v if v is not None else "")

        # Zeilenfarbe
        if is_excl:
            row_fill = FILL_EXCLUDE
        elif is_sens:
            row_fill = FILL_SENSIBEL
        elif is_res:
            row_fill = FILL_RESERVE
        elif is_leu:
            row_fill = FILL_LEUCHT
        else:
            row_fill = None

        if row_fill:
            for ci in range(1, len(COLS) + 1):
                ws.cell(row=ri, column=ci).fill = row_fill

        # Orange thema-Zelle bei unvollständiger erg (überschreibt Zeilenfarbe)
        if erg_warn:
            ws.cell(row=ri, column=2).fill = FILL_ERG_WARN

    # Spaltenbreiten
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "C2"   # themengebiet + thema sichtbar beim Scrollen
    ws.auto_filter.ref = f"A1:S{len(all_sorted) + 1}"

    # Statistik-Sheet
    ws2 = wb.create_sheet("Statistik")
    ws2.append(["Themengebiet", "Total", "Primary", "Reserve", "Exclude", "Sensibel", "Leuchtturm", "Erg-Lücke"])
    ws2["A1"].font = FONT_BOLD
    from collections import defaultdict
    by_g: dict = defaultdict(list)
    for item in canonical:
        by_g[item.get("themengebiet", "—")].append(item)
    totals = [0] * 7
    for g in sorted(by_g.keys()):
        gi = by_g[g]
        row = [
            g,
            len(gi),
            sum(1 for x in gi if x.get("tier") == "primary" and x.get("eignung") != "exclude"),
            sum(1 for x in gi if x.get("tier") == "reserve"),
            sum(1 for x in gi if x.get("eignung") == "exclude"),
            sum(1 for x in gi if x.get("sensibel")),
            sum(1 for x in gi if x.get("leuchtturm")),
            sum(1 for x in gi if erg_incomplete(x) and x.get("eignung") != "exclude"),
        ]
        ws2.append(row)
        for i, v in enumerate(row[1:], 0):
            totals[i] += v
    ws2.append(["GESAMT"] + totals)
    ws2.cell(row=ws2.max_row, column=1).font = FONT_BOLD
    ws2.column_dimensions["A"].width = 30

    wb.save(OUT)
    print(f"\n✓ {OUT.name} geschrieben: {len(all_sorted)} Zeilen, {n_orange} orange (erg-Lücke)")
    print(f"  Sortierung: themengebiet + thema alpha")
    print(f"  Freeze: C2 | AutoFilter: A1:R{len(all_sorted)+1}")


if __name__ == "__main__":
    main()
