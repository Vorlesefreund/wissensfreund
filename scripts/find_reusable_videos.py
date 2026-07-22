#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
find_reusable_videos.py — Suchmaschine fuer NACHNUTZBARE (re-hostbare) Videos.

Zweck (Wissensfreund):
    Fuer ausgewaehlte Leuchtturmthemen offen lizenzierte Videos finden, die wir
    LEGAL herunterladen und auf eigenem Server hosten duerfen (keine YouTube-
    Einbettung => keine Kinderdaten an Google). Ergebnis ist eine anklickbare
    Review-HTML-Datei: Du schaust die Kandidaten an und waehlst aus.

Rechts-Ampel je Quelle (KEINE Rechtsberatung, nur Heuristik):
    GRUEN  Wikimedia Commons / NASA  -> freie Lizenz bzw. Public Domain,
           re-hostbar (CC-BY: Namensnennung Pflicht!).
    GELB   Internet Archive / YouTube-CC -> Lizenz je Upload PRUEFEN. YouTube-CC
           bedeutet CC-BY *laut Uploader* -> Gefahr "License-Laundering"
           (Uploader besitzt Rechte evtl. gar nicht). Immer manuell verifizieren.

Quellen ohne API-Key (laufen sofort): Terra X (ZDF, CC-BY, DEUTSCH), USGS + NOAA
    (US-Behoerden, Public Domain), Wikimedia Commons, NASA, Internet Archive.
    Terra X/USGS/NOAA laufen ueber Commons-Kategorien (incategory-Filter).
Quellen mit Gratis-Key (Umgebungsvariablen):
    PEXELS_API_KEY   -> Pexels Videos (stummes Stock-B-Roll)   pexels.com/api
    PIXABAY_API_KEY  -> Pixabay Videos (stummes Stock-B-Roll)  pixabay.com/api/docs
    YT_API_KEY       -> YouTube CC-BY (breit, aber engl.+Risiko) ~100 Suchen/Tag
Stock (Pexels/Pixabay) ist meist STUMM -> ideal, um eigene deutsche Erzaehlung
    (Flash-Skript aus Klexikon + Flash-TTS, ~7 Cent/5 Min) drueberzulegen.

Ausgabe: Excel auf dem Desktop (video_kandidaten.xlsx) mit Vorschaubild,
    klickbaren Links, Farb-Ampel, Filter und leeren Spalten Verwenden?/Note/
    Kommentar zum Bewerten. Optional zusaetzlich --html und --json.

Beispiele:
    python scripts/find_reusable_videos.py --topics "Vulkan::volcano eruption,Wal::whale"
    python scripts/find_reusable_videos.py --topics-file leuchttuerme.txt
    python scripts/find_reusable_videos.py --topics "Vulkan::volcano" --html C:\\...\\v.html

Kurze/mehrdeutige dt. Woerter: Alias-Syntax "Anzeige::Suchbegriff" nutzen.
API-Keys werden auch aus .env (Repo-Wurzel) gelesen (echte Env-Vars haben Vorrang).
Vorauswahl: --min-sec/--max-sec (Laenge) + --top N (beste N je Thema).
Braucht openpyxl (+ Pillow fuer Vorschaubilder). Sonst nur Standardbibliothek.
"""

import argparse
import functools
import html
import json
import os
import re
import shutil
import subprocess
import sys
import time
import urllib.parse
import urllib.request

# Windows-Konsole auf UTF-8 zwingen (Projekt-Gotcha)
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass

USER_AGENT = "Wissensfreund-VideoFinder/1.0 (Bildungs-App; Kontakt: az@expansionssupport.de)"
TIMEOUT = 25


def _get_json(url):
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=TIMEOUT) as resp:
        return json.loads(resp.read().decode("utf-8", "replace"))


def classify_license(text):
    """Ampel fuer KOMMERZIELLE Nutzung (App ist bezahlpflichtig).
    ROT  = NC (nicht-kommerziell) oder ND (keine Bearbeitung) -> verboten.
    GRUEN= CC0/PD/CC-BY/CC-BY-SA -> nutzbar (BY/SA: Namensnennung/Weitergabe).
    GELB = unklar/keine Angabe -> pruefen."""
    t = (text or "").lower()
    if "-nc" in t or " nc " in t or "noncommercial" in t or "non-commercial" in t \
       or "-nd" in t or "noderiv" in t or "no-deriv" in t:
        return "red"
    if ("cc0" in t or "public domain" in t or "publicdomain" in t
            or "cc by" in t or "cc-by" in t or "creativecommons.org/licenses/by" in t):
        return "green"
    return "yellow"


def _fmt_secs(s):
    try:
        s = int(float(s))
    except (TypeError, ValueError):
        return ""
    return f"{s // 60}:{s % 60:02d}"


def _to_sec(s):
    """Rohsekunden als float oder None (fuer Laengenfilter)."""
    try:
        return float(s) if s not in (None, "") else None
    except (TypeError, ValueError):
        return None


def analyze_thumb(url):
    """Generische Bildanalyse des Vorschaubilds (kein Theme-Hardcoding):
    Anteil WARMER Toene (Feuer/Glut/Lava) und DUNKLER Toene (Rauch/Nacht) in %.
    Warm = R kraeftig und deutlich groesser als Blau (deckt rot/orange/gelb ab)."""
    try:
        import io
        from PIL import Image as PILImage
        req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
        raw = urllib.request.urlopen(req, timeout=15).read()
        im = PILImage.open(io.BytesIO(raw)).convert("RGB")
        im.thumbnail((100, 100))
        px = list(im.getdata())
    except Exception:
        return None, None
    if not px:
        return None, None
    warm = dark = 0
    for r, g, b in px:
        if r >= 120 and r >= g >= b and (r - b) >= 45:
            warm += 1
        if r + g + b <= 140:
            dark += 1
    n = len(px)
    return round(100 * warm / n, 1), round(100 * dark / n, 1)


def _res_label(w, h):
    """'3840x2160 4K' etc. + Pixelzahl (fuer Sortierung nach Aufloesung)."""
    try:
        w, h = int(w), int(h)
    except (TypeError, ValueError):
        return "", 0
    if not w or not h:
        return "", 0
    if w >= 3840 or h >= 2160:
        tag = "4K"
    elif w >= 1920 or h >= 1080:
        tag = "1080p"
    elif h >= 720:
        tag = "720p"
    else:
        tag = "SD"
    return f"{w}x{h} {tag}", w * h


# ---------------------------------------------------------------------------
# ffprobe: echte Laenge + Tonspur + Sprach-Tag (falls gesetzt)
# ---------------------------------------------------------------------------
FFPROBE = shutil.which("ffprobe")
DE_TAGS = {"de", "ger", "deu", "german"}
_GENERIC_TAGS = {"", "und", "unknown", "mis", "mul"}


def probe_media(url, timeout=30, retries=1):
    """Liefert {duration, has_audio, has_video, lang} oder None.
    Retry gegen kurzzeitige Server-Drosselung (viele ffprobe-Bursts)."""
    if not FFPROBE or not url:
        return None
    data = None
    for attempt in range(retries + 1):
        try:
            p = subprocess.run(
                [FFPROBE, "-v", "error", "-user_agent", USER_AGENT, "-show_entries",
                 "format=duration:stream=codec_type:stream_tags=language",
                 "-of", "json", url],
                capture_output=True, text=True, timeout=timeout)
            if p.returncode == 0:
                data = json.loads(p.stdout or "{}")
                break
        except Exception:
            pass
        if attempt < retries:
            time.sleep(0.5)
    if data is None:
        return None
    streams = data.get("streams", []) or []
    has_audio = any(s.get("codec_type") == "audio" for s in streams)
    has_video = any(s.get("codec_type") == "video" for s in streams)
    lang = ""
    for s in streams:
        if s.get("codec_type") == "audio":
            lang = ((s.get("tags") or {}).get("language") or "").lower()
            if lang:
                break
    if not lang:
        for s in streams:
            t = ((s.get("tags") or {}).get("language") or "").lower()
            if t:
                lang = t
                break
    dur = None
    try:
        dur = float((data.get("format") or {}).get("duration"))
    except (TypeError, ValueError):
        pass
    return {"duration": dur, "has_audio": has_audio, "has_video": has_video, "lang": lang}


def derive_language(probe, source):
    """(Anzeigetext, Rang) — Rang: 0=ideal (stumm/dt), 1=pruefen, 2=wohl fremd, 3=kein Video."""
    if probe is not None:
        if not probe["has_video"] and probe["has_audio"]:
            return ("AUDIO-Datei (kein Video)", 3)
        if not probe["has_audio"]:
            return ("ohne Ton — ideal", 0)
        lang = probe["lang"]
        if lang == "zxx":  # ISO: "kein sprachlicher Inhalt"
            return ("ohne Sprache (Tag zxx) — ideal", 0)
        if lang in DE_TAGS:
            return ("Deutsch (Tag) — ideal", 0)
        if lang and lang not in _GENERIC_TAGS:
            return (f"Ton, Tag: {lang} — pruefen (evtl. nur Umgebungston)", 1)
        return ("Ton vorhanden — Sprache pruefen", 1)
    # ohne Probe: grobe Quellen-Heuristik
    if source.startswith("NASA"):
        return ("meist Englisch — pruefen", 2)
    if source.startswith("YouTube"):
        return ("oft Englisch — pruefen", 1)
    if source.startswith("Internet"):
        return ("pruefen", 1)
    return ("unbekannt — pruefen", 1)


def _nasa_mp4(nasa_id):
    """Aus dem NASA-Asset-Manifest eine echte .mp4-URL holen (kleinste zuerst)."""
    try:
        d = _get_json(f"https://images-api.nasa.gov/asset/{urllib.parse.quote(nasa_id)}")
        hrefs = [i.get("href", "") for i in ((d.get("collection") or {}).get("items") or [])]
    except Exception:
        return ""
    mp4s = [h for h in hrefs if h.lower().endswith(".mp4")]
    if not mp4s:
        return ""
    # bevorzugt kleine Varianten (mobile/small) fuers schnelle Proben
    for key in ("~mobile", "small", "~small", "preview"):
        for h in mp4s:
            if key in h.lower():
                return h
    return mp4s[-1]


# ---------------------------------------------------------------------------
# Quelle 1: Wikimedia Commons  (freie Lizenzen, re-hostbar)  -- kein Key
# ---------------------------------------------------------------------------
def search_commons(term, limit, probe=True, category=None,
                   source_label="Wikimedia Commons",
                   note="CC-BY => Namensnennung Pflicht. Direkter Datei-Download erlaubt.",
                   lang_override=None, extra=None):
    search = f"filetype:video {term}"
    if category:  # nur Dateien dieser Commons-Kategorie (z.B. Terra X)
        search += f' incategory:"{category}"'
    if extra:     # zusaetzlicher Volltext-Anker (z.B. Behoerdenname statt Kategorie)
        search += f" {extra}"
    q = urllib.parse.urlencode({
        "action": "query", "format": "json", "generator": "search",
        "gsrsearch": search, "gsrnamespace": "6",
        "gsrlimit": str(limit),
        "prop": "imageinfo",
        "iiprop": "url|extmetadata|size|mime|mediatype|duration",
        "iiurlwidth": "160",  # erzwingt thumburl (Standbild) fuer Videos
    })
    url = "https://commons.wikimedia.org/w/api.php?" + q
    out = []
    try:
        data = _get_json(url)
    except Exception as e:
        return out, f"{source_label}-Fehler: {e}"
    pages = (data.get("query") or {}).get("pages") or {}
    for p in pages.values():
        ii = (p.get("imageinfo") or [{}])[0]
        meta = ii.get("extmetadata") or {}

        def m(key):
            v = meta.get(key) or {}
            return (v.get("value") or "").strip()

        # HTML aus Artist/Attribution grob entschaerfen
        artist = m("Artist") or m("Attribution")
        artist = _strip_tags(artist)
        lic = m("LicenseShortName") or m("License")
        file_url = ii.get("url", "")
        pr = probe_media(file_url) if probe else None
        lang_txt, lang_rank = derive_language(pr, source_label)
        if lang_override is not None:  # Quelle mit bekannter Sprache (z.B. Terra X = dt.)
            lang_txt, lang_rank = lang_override
        raw = (pr["duration"] if (pr and pr.get("duration")) else ii.get("duration"))
        dur_sec = _to_sec(raw)
        out.append({
            "source": source_label,
            "flag": classify_license(lic),
            "title": p.get("title", "").replace("File:", ""),
            "author": artist or "unbekannt",
            "license": lic or "siehe Dateiseite",
            "duration": _fmt_secs(raw),
            "dur_sec": dur_sec,
            "language": lang_txt,
            "lang_rank": lang_rank,
            "page": ii.get("descriptionurl", ""),
            "file": file_url,
            "thumb": ii.get("thumburl") or "",
            "note": note,
        })
    return out, None


# ---------------------------------------------------------------------------
# Quelle 7-9: Terra X (ZDF, CC-BY, DEUTSCH) + USGS/NOAA (US-Behoerde, PD)
#   -- alle drei ueber Commons-Kategorien, kein Key noetig
# ---------------------------------------------------------------------------
def search_terrax(term, limit, probe=True):
    return search_commons(
        term, limit, probe,
        category="Videos by Terra X",
        source_label="Terra X (ZDF, CC)",
        note=("ZDF/Terra X — CC-BY: Namensnennung 'ZDF/Terra X' + Lizenz + Link Pflicht. "
              "BY-SA-Varianten = Copyleft (euer Video muss dann auch BY-SA sein)."),
        lang_override=("Deutsch (ZDF/Terra X)", 0))


def search_usgs(term, limit, probe=True):
    # Die Kategorie "Videos by the USGS" enthaelt nur 6 Dateien -> Volltext-Anker
    # statt incategory (186 Treffer). Gleiches gilt fuer NOAA (1 vs. ~12.000).
    return search_commons(
        term, limit, probe, extra="USGS",
        source_label="USGS (PD)",
        note="US-Behoerde -> meist Public Domain. Im Zweifel Dateiseite pruefen.")


def search_noaa(term, limit, probe=True):
    return search_commons(
        term, limit, probe, extra="NOAA",
        source_label="NOAA (PD)",
        note="US-Behoerde -> meist Public Domain. Im Zweifel Dateiseite pruefen.")


def search_gailhampshire(term, limit, probe=True):
    """367 CC-BY-Tiervideos (Voegel, Insekten, Reptilien) — Nahaufnahmen, i.d.R. ohne Sprache."""
    return search_commons(
        term, limit, probe,
        category="Videos by Gailhampshire",
        source_label="Gailhampshire (CC-BY, Tiere)",
        note="CC-BY 2.0 — Namensnennung 'Gailhampshire' + Lizenz + Link Pflicht. "
             "Tiernahaufnahmen ohne Sprache: ideal fuer eigene Hoerspiel-Erzaehlung.")


def search_forstmeier(term, limit, probe=True):
    """333 CC-BY-Videos (Insekten/Heuschrecken) — Originalton = Tierstimmen."""
    return search_commons(
        term, limit, probe,
        category="Videos by Wolfgang Forstmeier",
        source_label="Forstmeier (CC-BY, Insekten)",
        note="CC-BY 3.0/4.0 — Namensnennung 'Wolfgang Forstmeier' + Lizenz + Link Pflicht. "
             "Originalton = Tierstimmen (Zirpen/Summen), keine Sprache.")


def search_esa(term, limit, probe=True):
    """ESA auf Commons: Kategorie 'Videos from ESA' hat 6 Dateien, Hubble 47.
    Zu duenn fuer eine eigene Quelle — Weltraum deckt die NASA-API ab.
    Achtung: NICHT per Volltext 'ESA OR Hubble' bauen — die OR-Gruppe haengt
    den Themenbegriff ab und liefert fuer JEDES Thema dieselben Teleskop-Clips."""
    return search_commons(
        term, limit, probe,
        category="Videos from ESA",
        source_label="ESA (CC BY-SA IGO)",
        note="ESA-Material meist CC BY-SA 3.0 IGO = Copyleft. Dateiseite pruefen.")


def _strip_tags(s):
    import re
    return re.sub(r"<[^>]+>", " ", s or "").strip()


# ---------------------------------------------------------------------------
# Quelle 2: NASA Image & Video Library (Public Domain)  -- kein Key
# ---------------------------------------------------------------------------
def search_nasa(term, limit, probe=True):
    q = urllib.parse.urlencode({"q": term, "media_type": "video"})
    url = "https://images-api.nasa.gov/search?" + q
    out = []
    try:
        data = _get_json(url)
    except Exception as e:
        return out, f"NASA-Fehler: {e}"
    items = ((data.get("collection") or {}).get("items") or [])[:limit]
    for it in items:
        d = (it.get("data") or [{}])[0]
        nid = d.get("nasa_id", "")
        links = it.get("links") or []
        thumb = next((l.get("href") for l in links if l.get("render") == "image"), "")
        mp4 = _nasa_mp4(nid)
        pr = probe_media(mp4) if (probe and mp4) else None
        lang_txt, lang_rank = derive_language(pr, "NASA")
        raw = pr["duration"] if (pr and pr.get("duration")) else None
        out.append({
            "source": "NASA",
            "flag": "green",
            "title": d.get("title", nid),
            "author": d.get("center", "NASA"),
            "license": "Public Domain (NASA media guidelines)",
            "duration": _fmt_secs(raw) if raw else "",
            "dur_sec": _to_sec(raw),
            "language": lang_txt,
            "lang_rank": lang_rank,
            "page": f"https://images.nasa.gov/details/{nid}",
            "file": mp4 or f"https://images-api.nasa.gov/asset/{urllib.parse.quote(nid)}",
            "thumb": thumb,
            "note": "Frei nutzbar; kein Endorsement suggerieren.",
        })
    return out, None


# ---------------------------------------------------------------------------
# Quelle 3: Internet Archive (gemischt: PD/CC + Unklares)  -- kein Key
# ---------------------------------------------------------------------------
def search_archive(term, limit, probe=True):
    q = urllib.parse.urlencode({
        "q": f'({term}) AND mediatype:movies',
        "rows": str(limit), "output": "json",
        "fl[]": "identifier",  # weitere fl unten manuell angehaengt
    })
    # mehrere fl[] Felder:
    url = ("https://archive.org/advancedsearch.php?" + q +
           "&fl[]=title&fl[]=creator&fl[]=licenseurl&fl[]=year")
    out = []
    try:
        data = _get_json(url)
    except Exception as e:
        return out, f"Archive-Fehler: {e}"
    docs = ((data.get("response") or {}).get("docs") or [])[:limit]
    for doc in docs:
        lic = doc.get("licenseurl") or ""
        flag = classify_license(lic) if lic else "yellow"
        ident = doc.get("identifier", "")
        lang_txt, lang_rank = derive_language(None, "Internet Archive")
        out.append({
            "source": "Internet Archive",
            "flag": flag,
            "title": doc.get("title", ident),
            "author": _asstr(doc.get("creator")) or "unbekannt",
            "license": lic or "KEINE Lizenz angegeben — pruefen!",
            "duration": "",
            "language": lang_txt,
            "lang_rank": lang_rank,
            "page": f"https://archive.org/details/{ident}",
            "file": f"https://archive.org/download/{ident}",
            "thumb": f"https://archive.org/services/img/{ident}",
            "note": "Lizenz je Upload PRUEFEN (viele Uploads ohne saubere Rechte).",
        })
    return out, None


def _asstr(v):
    if isinstance(v, list):
        return ", ".join(str(x) for x in v)
    return str(v) if v else ""


# ---------------------------------------------------------------------------
# Quelle 4: YouTube Data API v3, CC-BY-Filter (optional, Key noetig)
# ---------------------------------------------------------------------------
def search_youtube_cc(term, limit, lang, api_key, probe=False):
    out = []
    q = urllib.parse.urlencode({
        "part": "snippet", "q": term, "type": "video",
        "videoLicense": "creativeCommon", "safeSearch": "strict",
        "videoEmbeddable": "true", "maxResults": str(min(limit, 25)),
        "relevanceLanguage": lang, "key": api_key,
    })
    try:
        data = _get_json("https://www.googleapis.com/youtube/v3/search?" + q)
    except Exception as e:
        return out, f"YouTube-Fehler: {e}"
    ids = [it["id"]["videoId"] for it in data.get("items", []) if it.get("id", {}).get("videoId")]
    durations = {}
    if ids:
        try:
            q2 = urllib.parse.urlencode({
                "part": "contentDetails", "id": ",".join(ids), "key": api_key})
            det = _get_json("https://www.googleapis.com/youtube/v3/videos?" + q2)
            for v in det.get("items", []):
                durations[v["id"]] = _iso_dur(v.get("contentDetails", {}).get("duration", ""))
        except Exception:
            pass
    for it in data.get("items", []):
        vid = it.get("id", {}).get("videoId")
        if not vid:
            continue
        sn = it.get("snippet", {})
        lang_txt, lang_rank = derive_language(None, "YouTube (CC-BY)")
        out.append({
            "source": "YouTube (CC-BY)",
            "flag": "yellow",
            "title": html.unescape(sn.get("title", vid)),
            "author": sn.get("channelTitle", ""),
            "license": "CC-BY (laut Uploader)",
            "duration": durations.get(vid, ""),
            "language": lang_txt,
            "lang_rank": lang_rank,
            "page": f"https://www.youtube.com/watch?v={vid}",
            "file": f"https://www.youtube.com/watch?v={vid}",
            "thumb": (sn.get("thumbnails", {}).get("medium", {}) or {}).get("url", ""),
            "note": "License-Laundering-Risiko: pruefen, ob Uploader wirklich Rechteinhaber ist.",
        })
    return out, None


# Oeffentlich-rechtliche Kanaele: duerfen laut Telemedienwerbeverbot (Medienstaats-
# vertrag) auf YouTube KEINE Werbung/Sponsoring schalten -> rechtlich werbefrei.
# WICHTIG: nur konkrete Kinder-/Wissensformate — KEINE pauschalen Senderkuerzel
# (zdf/ard/wdr), die auch Erwachsenen-/Satirekanaele derselben Anstalt treffen.
OER_CHANNELS = [
    "sendung mit der maus", "die maus", "der elefant",
    "die sendung mit dem elefanten", "checker", "logo!", "terra x",
    "kika", "zdftivi", "planet schule", "wissen macht ah",
    "löwenzahn", "loewenzahn", "willi wills wissen", "pur+", "neuneinhalb",
    "sesamstraße", "sesamstrasse", "anna und die", "paula und die",
    "swr kindernetz", "tigerenten", "schau in meine welt", "quarks",
    "die pfefferkörner", "erklär mir die welt", "kikaninchen",
]


def _channel_allowed(title, allow):
    """Kanalname gegen Allowlist. Kurze Kuerzel (ard/zdf/wdr) nur als ganzes Wort."""
    t = (title or "").lower()
    for frag in allow:
        f = frag.strip().lower()
        if not f:
            continue
        if len(f) <= 4 and f.isalpha():
            if re.search(r"\b" + re.escape(f) + r"\b", t):
                return True
        elif f in t:
            return True
    return False


# ---------------------------------------------------------------------------
# Kanal-Register fuer die GEZIELTE Suche IN vertrauenswuerdigen Kanaelen.
# tier "oer"  = oeffentlich-rechtlich -> Telemedienwerbeverbot -> WERBEFREI garantiert
# tier "komm" = kommerzieller Kinderkanal -> WERBUNG MOEGLICH (nicht abschaltbar!)
# ---------------------------------------------------------------------------
YT_CHANNEL_REGISTRY = [
    ("Die Sendung mit der Maus", "oer"),
    ("Der Elefant", "oer"),
    ("Checker Welt", "oer"),
    ("Terra X plus", "oer"),
    ("logo!", "oer"),
    ("Löwenzahn", "oer"),
    ("Wissen macht Ah!", "oer"),
    ("Anna und die wilde Tiere", "oer"),
    ("planet schule", "oer"),
    ("KiKA", "oer"),
    ("pur+", "oer"),
    ("Woozle Goozle", "komm"),
    ("TOGGO", "komm"),
    ("Lernen mit Glapi", "komm"),
]
AD_LABEL = {"oer": "werbefrei (oeffentlich-rechtlich)",
            "komm": "WERBUNG MOEGLICH (kommerziell)"}


def _channel_cache_path():
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), "_yt_channel_cache.json")


def resolve_channel_ids(names, api_key):
    """Kanalnamen -> channelId. Ergebnis wird gecacht (spart API-Kontingent:
    jede Aufloesung kostet 100 Einheiten)."""
    path = _channel_cache_path()
    cache = {}
    if os.path.isfile(path):
        try:
            with open(path, encoding="utf-8") as f:
                cache = json.load(f)
        except Exception:
            cache = {}
    missing = [n for n in names if n not in cache]
    for name in missing:
        try:
            q = urllib.parse.urlencode({"part": "snippet", "type": "channel",
                                        "q": name, "maxResults": "1", "key": api_key})
            d = _get_json("https://www.googleapis.com/youtube/v3/search?" + q)
            items = d.get("items", [])
            cache[name] = items[0]["snippet"]["channelId"] if items else ""
            print(f"    Kanal aufgeloest: {name} -> {cache[name] or 'NICHT GEFUNDEN'}")
        except Exception as e:
            print(f"    Kanal-Aufloesung fehlgeschlagen ({name}): {e}")
            cache[name] = ""
    if missing:
        try:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(cache, f, ensure_ascii=False, indent=2)
        except Exception:
            pass
    return {n: cache.get(n, "") for n in names}


def _katalog_dir():
    d = os.path.join(os.path.dirname(os.path.abspath(__file__)), "_yt_katalog")
    os.makedirs(d, exist_ok=True)
    return d


def build_channel_catalog(name, tier, cid, api_key, max_videos=600, refresh=False):
    """Kanal-Uploads EINMAL holen, danach alle Themensuchen offline.

    search.list ist auf 100 Anfragen/Tag gedeckelt — playlistItems.list NICHT
    (1 Einheit je 50 Videos). Ein Katalog kostet ~25 Einheiten und bedient
    beliebig viele Themen ohne weiteren Verbrauch."""
    path = os.path.join(_katalog_dir(), f"{cid}.json")
    if not refresh and os.path.isfile(path):
        try:
            with open(path, encoding="utf-8") as f:
                cat = json.load(f)
            # gedeckelter Katalog + jetzt mehr Tiefe verlangt -> neu holen.
            # Altkatalog ohne Tiefen-Angabe: Deckel unbekannt -> Videozahl annehmen.
            tiefe = cat.get("max_videos", len(cat.get("videos", [])))
            deckel = cat.get("gedeckelt", len(cat.get("videos", [])) >= tiefe)
            zu_flach = deckel and tiefe < max_videos
            if cat.get("videos") and not zu_flach:
                return cat, None
        except Exception:
            pass
    try:
        q = urllib.parse.urlencode({"part": "contentDetails", "id": cid, "key": api_key})
        d = _get_json("https://www.googleapis.com/youtube/v3/channels?" + q)
        items = d.get("items", [])
        if not items:
            return None, f"{name}: Kanal nicht gefunden"
        pl = items[0]["contentDetails"]["relatedPlaylists"]["uploads"]
    except Exception as e:
        return None, f"{name}: {e}"

    vids, token = [], None
    while len(vids) < max_videos:
        p = {"part": "snippet", "playlistId": pl, "maxResults": "50", "key": api_key}
        if token:
            p["pageToken"] = token
        try:
            d = _get_json("https://www.googleapis.com/youtube/v3/playlistItems?"
                          + urllib.parse.urlencode(p))
        except Exception as e:
            if not vids:
                return None, f"{name}: {e}"
            break
        for it in d.get("items", []):
            sn = it.get("snippet", {})
            vid = (sn.get("resourceId") or {}).get("videoId")
            if vid:
                vids.append({"id": vid, "title": html.unescape(sn.get("title", "")),
                             "description": sn.get("description", "")[:600],
                             "published": sn.get("publishedAt", ""),
                             "thumb": (sn.get("thumbnails", {}).get("medium", {}) or {}).get("url", "")})
        token = d.get("nextPageToken")
        if not token:
            break

    for i in range(0, len(vids), 50):                       # Dauer nachziehen
        chunk = vids[i:i + 50]
        try:
            q2 = urllib.parse.urlencode({"part": "contentDetails",
                                         "id": ",".join(v["id"] for v in chunk),
                                         "key": api_key})
            det = _get_json("https://www.googleapis.com/youtube/v3/videos?" + q2)
            secs = {v["id"]: _iso_sec((v.get("contentDetails") or {}).get("duration", ""))
                    for v in det.get("items", [])}
            for v in chunk:
                v["sec"] = secs.get(v["id"])
        except Exception:
            pass

    cat = {"name": name, "tier": tier, "channel_id": cid,
           "max_videos": max_videos, "gedeckelt": len(vids) >= max_videos,
           "videos": vids}
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(cat, f, ensure_ascii=False, indent=1)
    except Exception:
        pass
    return cat, None


def search_catalog(term, channels, api_key, per_channel=8, refresh=False, katalog_max=600):
    """Offline-Themensuche im lokalen Kanalkatalog (kostet nach Aufbau NICHTS)."""
    out = []
    for name, tier, cid in channels:
        if not cid:
            continue
        cat, err = build_channel_catalog(name, tier, cid, api_key,
                                         max_videos=katalog_max, refresh=refresh)
        if err:
            print(f"    {name}: {err}")
            continue
        hits = [v for v in cat["videos"] if _title_matches(term, v)]
        hits.sort(key=lambda v: v.get("published") or "", reverse=True)   # neueste zuerst
        treffer = len(hits)
        hits = hits[:per_channel]
        print(f"    {name}: {treffer} von {len(cat['videos'])} Videos passend"
              + (f" (zeige {per_channel})" if treffer > per_channel else ""))
        for v in hits:
            sec = v.get("sec")
            out.append({
                "source": f"YT {name}",
                "flag": "embed",
                "title": v["title"],
                "author": name,
                "license": f"Einbetten · {AD_LABEL[tier]}",
                "duration": _fmt_secs(sec) if sec else "", "dur_sec": sec,
                "language": "Deutsch (Kanal)" if tier == "oer" else "Deutsch (Kanal, komm.)",
                "lang_rank": 0 if tier == "oer" else 1,
                "page": f"https://www.youtube.com/watch?v={v['id']}",
                "file": f"https://www.youtube-nocookie.com/embed/{v['id']}",
                "thumb": v.get("thumb", ""),
                "note": (f"{AD_LABEL[tier]}. Einbetten (kein Re-Host), Eltern-Opt-in noetig. "
                         "Zu lang? start=/end= im Embed nutzen (schneidet KEINE Werbung!)."),
            })
    return out, None


_DE_ENDUNGEN = r"(?:e|en|es|s|er|n)?"


@functools.lru_cache(maxsize=256)
def _term_pattern(word):
    """Deutsche Komposita ohne blinden Substring-Treffer.

    Immer: eigenstaendiges Wort samt Pluralendung (Wal, Wale) und Kompositum-
    ENDE (Buckelwal, Blauwale). NICHT: Gewalt, Anwalt, Qualle.
    Kompositum-ANFANG (Vulkanausbruch) nur ab 5 Zeichen — bei kurzen Woertern
    waeren zu viele Fremdwoerter blosse Praefix-Zufaelle (Wal->Wald, Walzer)."""
    w = re.escape(word)
    arme = [r"(?:\w+)?" + w + _DE_ENDUNGEN + r"\b"]
    if len(word) >= 5:
        arme.append(r"\b" + w)
    return re.compile("|".join(arme))


def _title_matches(term, sn):
    """Kanal-Suche fuellt mit Themenfremdem auf -> Begriff muss im TITEL stehen.
    Beschreibungen sind Boilerplate-lastig und erzeugen zu viel Rauschen."""
    hay = (sn.get("title", "") or "").lower()
    words = [w for w in re.split(r"\W+", (term or "").lower()) if len(w) >= 3]
    if not words:
        return True
    return any(_term_pattern(w).search(hay) for w in words)


def search_youtube_channels(term, api_key, channels, per_channel=5, lang="de", strict=True):
    """GEZIELTE Suche IN vertrauenswuerdigen Kanaelen (statt generisch + wegfiltern).
    channels: Liste (name, tier, channel_id)."""
    out = []
    all_ids = []
    staged = []
    for name, tier, cid in channels:
        if not cid:
            continue
        try:
            q = urllib.parse.urlencode({
                "part": "snippet", "type": "video", "channelId": cid, "q": term,
                "maxResults": str(min(per_channel, 25)), "safeSearch": "strict",
                "videoEmbeddable": "true", "key": api_key})
            d = _get_json("https://www.googleapis.com/youtube/v3/search?" + q)
        except Exception as e:
            print(f"    {name}: Fehler {e}")
            continue
        kept = off = 0
        for it in d.get("items", []):
            vid = (it.get("id") or {}).get("videoId")
            if not vid:
                continue
            sn = it.get("snippet", {})
            if strict and not _title_matches(term, sn):
                off += 1          # Kanal hat zum Thema nichts -> YT fuellt fremd auf
                continue
            staged.append((vid, sn, name, tier))
            all_ids.append(vid)
            kept += 1
        if kept or off:
            print(f"    {name}: {kept} passend" + (f", {off} themenfremd verworfen" if off else ""))
    # Dauer in einem Rutsch (1 Einheit pro 50 IDs)
    durs = {}
    for i in range(0, len(all_ids), 50):
        try:
            q2 = urllib.parse.urlencode({"part": "contentDetails",
                                         "id": ",".join(all_ids[i:i + 50]), "key": api_key})
            det = _get_json("https://www.googleapis.com/youtube/v3/videos?" + q2)
            for v in det.get("items", []):
                iso = (v.get("contentDetails") or {}).get("duration", "")
                durs[v["id"]] = _iso_sec(iso)
        except Exception:
            pass
    for vid, sn, name, tier in staged:
        sec = durs.get(vid)
        out.append({
            "source": f"YT {name}",
            "flag": "embed",
            "title": html.unescape(sn.get("title", vid)),
            "author": sn.get("channelTitle", name),
            "license": f"Einbetten · {AD_LABEL[tier]}",
            "duration": _fmt_secs(sec) if sec else "", "dur_sec": sec,
            "language": "Deutsch (Kanal)" if tier == "oer" else "Deutsch (Kanal, komm.)",
            "lang_rank": 0 if tier == "oer" else 1,
            "page": f"https://www.youtube.com/watch?v={vid}",
            "file": f"https://www.youtube-nocookie.com/embed/{vid}",
            "thumb": (sn.get("thumbnails", {}).get("medium", {}) or {}).get("url", ""),
            "note": (f"{AD_LABEL[tier]}. Einbetten (kein Re-Host), Eltern-Opt-in noetig. "
                     "Zu lang? start=/end= im Embed nutzen (schneidet KEINE Werbung!)."),
        })
    return out, None


def search_youtube_kids(term, limit, lang, api_key, allow_channels=None, require_kids=True):
    """YouTube-Videos zum EINBETTEN (Opt-in, kein Re-Host), NUR 'made for kids'.
    Kein CC-Filter (Einbetten braucht keine freie Lizenz). Filtert per status.madeForKids
    und optional auf eine Kanal-Allowlist (z.B. oeffentlich-rechtlich = werbefrei)."""
    out = []
    q = urllib.parse.urlencode({
        "part": "snippet", "q": term, "type": "video",
        "safeSearch": "strict", "videoEmbeddable": "true",
        "maxResults": str(min(limit, 50)),
        "relevanceLanguage": lang, "key": api_key,
    })
    try:
        data = _get_json("https://www.googleapis.com/youtube/v3/search?" + q)
    except Exception as e:
        return out, f"YouTube-Kids-Fehler: {e}"
    snip = {it["id"]["videoId"]: it.get("snippet", {})
            for it in data.get("items", []) if it.get("id", {}).get("videoId")}
    ids = list(snip.keys())
    if not ids:
        return out, None
    meta = {}
    try:
        q2 = urllib.parse.urlencode({"part": "contentDetails,status",
                                     "id": ",".join(ids), "key": api_key})
        det = _get_json("https://www.googleapis.com/youtube/v3/videos?" + q2)
        for v in det.get("items", []):
            iso = (v.get("contentDetails") or {}).get("duration", "")
            meta[v["id"]] = {
                "dur": _iso_dur(iso),
                "sec": _iso_sec(iso),
                "kids": bool((v.get("status") or {}).get("madeForKids", False)),
            }
    except Exception:
        pass
    n_kids = n_chan = 0
    for vid in ids:
        if require_kids and not meta.get(vid, {}).get("kids"):
            continue                            # nur 'made for kids'
        n_kids += 1
        sn = snip.get(vid, {})
        chan = sn.get("channelTitle", "")
        if allow_channels and not _channel_allowed(chan, allow_channels):
            continue                            # nur erlaubte (werbefreie) Kanaele
        n_chan += 1
        out.append({
            "source": "YouTube Kids (Embed)",
            "flag": "embed",
            "title": html.unescape(sn.get("title", vid)),
            "author": chan,
            "license": "Einbetten (kein Re-Host); Video bleibt bei YouTube",
            "duration": meta[vid].get("dur", ""), "dur_sec": meta[vid].get("sec"),
            "language": "YouTube-Sprache pruefen", "lang_rank": 1,
            "page": f"https://www.youtube.com/watch?v={vid}",
            "file": f"https://www.youtube-nocookie.com/embed/{vid}",
            "thumb": (sn.get("thumbnails", {}).get("medium", {}) or {}).get("url", ""),
            "note": ("NUR 'made for kids'. Einbetten via nocookie-Player; Eltern-Opt-in "
                     "noetig (Tracking an Google). Player abschotten (kein rel/Klickthrough)."),
        })
    print(f"    YouTube-Trichter: {len(ids)} Treffer -> {n_kids} nach Kids-Filter "
          f"-> {n_chan} nach Kanal-Filter")
    return out, None


def _iso_sec(iso):
    """ISO-8601-Dauer (PT#H#M#S) -> Sekunden oder None."""
    m = re.match(r"PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?", iso or "")
    if not m:
        return None
    h, mi, s = (int(x) if x else 0 for x in m.groups())
    return h * 3600 + mi * 60 + s


def _iso_dur(iso):
    t = _iso_sec(iso)
    return _fmt_secs(t) if t is not None else ""


# ---------------------------------------------------------------------------
# Quelle 5+6: Stock-Video (Pexels/Pixabay) — stumm, freie kommerzielle Lizenz
#   -> ideal: eigene deutsche Erzaehlung drueberlegen. Key ueber Env noetig.
# ---------------------------------------------------------------------------
STOCK_LANG = ("stumm/Stock — ideal (eigene dt. Stimme)", 0)
STOCK_NOTE = "Stock-Clip, i.d.R. ohne Sprache -> eigene deutsche Erzaehlung drueberlegen."


def search_pexels(term, limit, api_key):
    out = []
    q = urllib.parse.urlencode({"query": term, "per_page": min(limit, 80),
                                "locale": "de-DE", "orientation": "landscape"})
    req = urllib.request.Request("https://api.pexels.com/videos/search?" + q,
                                 headers={"User-Agent": USER_AGENT, "Authorization": api_key})
    try:
        with urllib.request.urlopen(req, timeout=TIMEOUT) as r:
            data = json.loads(r.read().decode("utf-8", "replace"))
    except Exception as e:
        return out, f"Pexels-Fehler: {e}"
    for v in data.get("videos", [])[:limit]:
        mp4 = ""
        for f in v.get("video_files", []):
            if f.get("file_type") == "video/mp4":
                mp4 = f.get("link", "")
                if f.get("quality") == "hd":
                    break
        slug = (v.get("url", "").rstrip("/").split("/") or [""])[-1]
        res, res_px = _res_label(v.get("width"), v.get("height"))
        out.append({
            "source": "Pexels (Stock)", "flag": "green",
            "title": (slug.replace("-", " ").title() if slug else f"pexels-{v.get('id')}"),
            "author": (v.get("user") or {}).get("name", "Pexels"),
            "license": "Pexels License (frei, kommerziell, ohne Namensnennung)",
            "duration": _fmt_secs(v.get("duration")), "dur_sec": _to_sec(v.get("duration")),
            "resolution": res, "res_px": res_px, "popularity": 0,
            "language": STOCK_LANG[0], "lang_rank": STOCK_LANG[1],
            "page": v.get("url", ""), "file": mp4 or v.get("url", ""),
            "thumb": v.get("image", ""), "note": STOCK_NOTE,
        })
    return out, None


def search_pixabay(term, limit, api_key):
    out = []
    per = max(3, min(limit, 200))
    q = urllib.parse.urlencode({"key": api_key, "q": term, "per_page": per,
                                "lang": "de", "safesearch": "true"})
    try:
        data = _get_json("https://pixabay.com/api/videos/?" + q)
    except Exception as e:
        return out, f"Pixabay-Fehler: {e}"
    for h in data.get("hits", [])[:limit]:
        vids = h.get("videos", {}) or {}
        pick = vids.get("large") or vids.get("medium") or vids.get("small") or {}
        res, res_px = _res_label(pick.get("width"), pick.get("height"))
        out.append({
            "source": "Pixabay (Stock)", "flag": "green",
            "title": (h.get("tags") or f"pixabay-{h.get('id')}"),
            "author": h.get("user", "Pixabay"),
            "license": "Pixabay Content License (frei, kommerziell, ohne Namensnennung)",
            "duration": _fmt_secs(h.get("duration")), "dur_sec": _to_sec(h.get("duration")),
            "resolution": res, "res_px": res_px, "popularity": int(h.get("views", 0) or 0),
            "language": STOCK_LANG[0], "lang_rank": STOCK_LANG[1],
            "page": h.get("pageURL", ""), "file": pick.get("url", ""),
            "thumb": pick.get("thumbnail", ""), "note": STOCK_NOTE,
        })
    return out, None


# ---------------------------------------------------------------------------
# HTML-Report
# ---------------------------------------------------------------------------
FLAG_LABEL = {
    "green": ("#1a7f37", "RE-HOSTBAR"),
    "yellow": ("#9a6700", "LIZENZ PRUEFEN"),
    "red": ("#b42318", "NICHT NUTZBAR (NC/ND)"),
    "embed": ("#1f6feb", "EINBETTEN (Opt-in)"),
}


def build_html(results_by_topic, path):
    """Kontaktbogen: klickbares Thumbnail-Raster zum schnellen visuellen Sichten."""
    parts = ["""<!doctype html><meta charset="utf-8">
<title>Video-Galerie — Wissensfreund</title>
<style>
 body{font:14px/1.5 system-ui,Segoe UI,sans-serif;max-width:1500px;margin:20px auto;padding:0 16px;color:#1c1c1c}
 h1{font-size:22px} h2{margin-top:34px;border-bottom:2px solid #ddd;padding-bottom:6px}
 .grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(240px,1fr));gap:14px;margin-top:12px}
 .card{border:1px solid #e2e2e2;border-radius:10px;overflow:hidden;background:#fff}
 .card a.thumb{display:block;position:relative}
 .card img{width:100%;height:150px;object-fit:cover;background:#eee;display:block}
 .dur{position:absolute;right:6px;bottom:6px;background:rgba(0,0,0,.75);color:#fff;font-size:12px;padding:1px 6px;border-radius:4px}
 .body{padding:8px 10px}
 .title{font-size:13px;font-weight:600;margin:0 0 4px;max-height:2.6em;overflow:hidden}
 .badge{display:inline-block;font-size:10px;font-weight:700;color:#fff;padding:1px 7px;border-radius:20px}
 .sub{font-size:11px;color:#666;margin-top:3px}
 .dl{font-size:11px}
 .legend{font-size:13px;background:#f7f7f7;border-radius:8px;padding:10px 14px;margin:14px 0}
</style>
<h1>Video-Galerie — visuell sichten &amp; picken</h1>
<div class="legend"><b style="color:#1a7f37">GRUEN</b> re-hostbar (CC-BY: Namensnennung!) &nbsp;
 <b style="color:#9a6700">GELB</b> Lizenz pruefen &nbsp; <b style="color:#b42318">ROT</b> NC/ND — nicht nutzbar &nbsp;
 <b style="color:#1f6feb">BLAU</b> Einbetten (Opt-in, Tracking; kein Re-Host).
 <br>Klick aufs Bild = Videoseite ansehen. Auswahl/Kommentar in der Excel. Kein Re-Host ohne verifizierte Lizenz.</div>
"""]
    for topic, cands in results_by_topic.items():
        parts.append(f"<h2>{html.escape(topic)} "
                     f"<span class='sub'>({len(cands)} Kandidaten)</span></h2>")
        if not cands:
            parts.append("<p><i>Keine Treffer.</i></p>")
            continue
        parts.append('<div class="grid">')
        for c in cands:
            color, label = FLAG_LABEL.get(c["flag"], ("#555", "?"))
            thumb = html.escape(c["thumb"]) if c["thumb"] else ""
            page = html.escape(c["page"])
            dur = f'<span class="dur">{html.escape(c["duration"])}</span>' if c["duration"] else ""
            img = f"<img src='{thumb}' loading='lazy' alt=''>" if thumb else "<img alt=''>"
            extra = " · ".join(x for x in (c.get("resolution", ""), c.get("language", "")) if x)
            parts.append(f"""<div class="card">
 <a class="thumb" href="{page}" target="_blank">{img}{dur}</a>
 <div class="body">
   <div class="title"><a href="{page}" target="_blank">{html.escape(c['title'])}</a></div>
   <span class="badge" style="background:{color}">{label}</span>
   <span class="sub">{html.escape(c['source'])}</span>
   <div class="sub">{html.escape(extra)}</div>
   <div class="sub dl"><a href="{html.escape(c['file'])}" target="_blank">Download</a> · {html.escape(c['license'][:28])}</div>
 </div></div>""")
        parts.append('</div>')
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(parts))


# ---------------------------------------------------------------------------
# Excel-Report (klickbar + kommentierbar)  -- Hauptausgabe
# ---------------------------------------------------------------------------
XLSX_FILLS = {"green": "C6EFCE", "yellow": "FFEB9C", "red": "FFC7CE", "embed": "CFE2FF"}
XLSX_AMPEL = {"green": "GRUEN – re-hostbar", "yellow": "GELB – pruefen",
              "red": "ROT – NC/ND", "embed": "BLAU – Einbetten (Opt-in)"}


def build_xlsx(results_by_topic, path, embed_thumbs=True):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    thumbs_ok = embed_thumbs
    if thumbs_ok:
        try:
            from openpyxl.drawing.image import Image as XLImage
            from PIL import Image as PILImage
        except Exception:
            thumbs_ok = False

    wb = Workbook()
    ws = wb.active
    ws.title = "Video-Kandidaten"
    link_font = Font(color="0563C1", underline="single")

    headers = ["Thema", "Vorschau", "Ampel", "Titel (Klick)", "Dauer", "Aufloesung",
               "Sprache", "Quelle", "Autor/Kanal", "Lizenz", "Download (Klick)", "Hinweis",
               "Verwenden?", "Note 1-5", "Kommentar"]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="404040")
        cell.alignment = Alignment(vertical="center")

    ideal_fill = PatternFill("solid", fgColor="D9F2D0")  # zarte Hervorhebung "ideal"
    row = 2
    for topic, cands in results_by_topic.items():
        for c in cands:
            ws.cell(row, 1, topic)
            acell = ws.cell(row, 3, XLSX_AMPEL.get(c["flag"], "?"))
            fill = XLSX_FILLS.get(c["flag"])
            if fill:
                acell.fill = PatternFill("solid", fgColor=fill)
            tcell = ws.cell(row, 4, c["title"])
            if c["page"]:
                tcell.hyperlink = c["page"]
                tcell.font = link_font
            ws.cell(row, 5, c["duration"])
            ws.cell(row, 6, c.get("resolution", ""))
            scell = ws.cell(row, 7, c.get("language", ""))
            if c.get("lang_rank", 1) == 0:
                scell.fill = ideal_fill
            scell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row, 8, c["source"])
            ws.cell(row, 9, c["author"])
            ws.cell(row, 10, c["license"])
            dcell = ws.cell(row, 11, "Datei oeffnen")
            if c["file"]:
                dcell.hyperlink = c["file"]
                dcell.font = link_font
            ncell = ws.cell(row, 12, c["note"])
            ncell.alignment = Alignment(wrap_text=True, vertical="top")
            tcell.alignment = Alignment(wrap_text=True, vertical="top")

            if thumbs_ok and c.get("thumb"):
                try:
                    req = urllib.request.Request(c["thumb"], headers={"User-Agent": USER_AGENT})
                    raw = urllib.request.urlopen(req, timeout=15).read()
                    im = PILImage.open(io.BytesIO(raw)).convert("RGB")
                    im.thumbnail((160, 90))
                    buf = io.BytesIO()
                    im.save(buf, format="PNG")
                    buf.seek(0)
                    xi = XLImage(buf)
                    ws.add_image(xi, f"B{row}")
                    ws.row_dimensions[row].height = 72
                except Exception:
                    pass
            row += 1

    widths = [16, 24, 20, 42, 8, 14, 24, 16, 18, 22, 15, 32, 13, 9, 32]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    if row > 2:
        last = get_column_letter(len(headers))
        ws.auto_filter.ref = f"A1:{last}{row - 1}"
        dv = DataValidation(type="list", formula1='"ja,nein,vielleicht"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"M2:M{row - 1}")  # Spalte "Verwenden?"
    wb.save(path)


def load_dotenv():
    """Minimaler .env-Loader (kein Paket noetig). Ueberschreibt keine echten Env-Vars."""
    here = os.path.dirname(os.path.abspath(__file__))
    for p in (".env", os.path.join(here, ".env"), os.path.join(here, "..", ".env")):
        if os.path.isfile(p):
            try:
                with open(p, encoding="utf-8") as f:
                    for line in f:
                        s = line.strip()
                        if not s or s.startswith("#") or "=" not in s:
                            continue
                        k, v = s.split("=", 1)
                        k = k.strip()
                        v = v.strip().strip('"').strip("'")
                        if k:
                            os.environ.setdefault(k, v)
            except Exception:
                pass
            return p
    return None


# ---------------------------------------------------------------------------
def main():
    load_dotenv()
    ap = argparse.ArgumentParser(description="Finde re-hostbare Videos fuer Themen.")
    g = ap.add_mutually_exclusive_group(required=True)
    g.add_argument("--topics", help="Kommagetrennt, z.B. \"Vulkan,Wal,Dinosaurier\"")
    g.add_argument("--topics-file", help="Textdatei, ein Thema pro Zeile")
    ap.add_argument("--lang", default="de", help="Sprachpraeferenz fuer YouTube (default de)")
    ap.add_argument("--channels", default="oer",
                    help="YouTube-Kanalfilter: 'oer' = nur oeffentlich-rechtliche "
                         "(gesetzlich werbefrei, Default) | 'all' = kein Filter | "
                         "eigene Liste kommagetrennt")
    ap.add_argument("--tier", default="oer", choices=["oer", "beide"],
                    help="Quelle ytchannels: 'oer' nur werbefreie oeffentlich-rechtliche "
                         "(Default) | 'beide' zusaetzlich kommerzielle Kinderkanaele "
                         "(WERBUNG MOEGLICH)")
    ap.add_argument("--per-channel", type=int, default=5,
                    help="Treffer je Kanal bei ytchannels (default 5)")
    ap.add_argument("--katalog-refresh", action="store_true",
                    help="ytkatalog: Kanalkataloge neu holen (sonst lokaler Cache)")
    ap.add_argument("--katalog-max", type=int, default=600,
                    help="ytkatalog: max. Videos je Kanal (default 600)")
    ap.add_argument("--locker", action="store_true",
                    help="ytchannels: Titel-Relevanzfilter AUS (mehr Treffer, mehr Themenfremdes)")
    ap.add_argument("--max", type=int, default=6, help="Kandidaten pro Quelle/Thema (default 6)")
    ap.add_argument("--min-sec", type=int, default=0, help="Mindestlaenge in Sekunden (0=aus)")
    ap.add_argument("--max-sec", type=int, default=0, help="Maximallaenge in Sekunden (0=aus)")
    ap.add_argument("--top", type=int, default=0,
                    help="Nur die besten N Kandidaten je Thema behalten (0=alle)")
    ap.add_argument("--sort", default="relevanz",
                    choices=["relevanz", "aufloesung", "beliebtheit", "warm"],
                    help="Reihung innerhalb Lizenz/Sprache: relevanz (default), "
                         "aufloesung (4K zuerst), beliebtheit (Pixabay-Views), "
                         "warm (Anteil Feuer-/Glut-Toene im Bild)")
    ap.add_argument("--sources",
                    default="terrax,pexels,pixabay,usgs,noaa,commons,nasa,archive,youtube",
                    help="Auswahl: terrax,gailhampshire,forstmeier,esa,pexels,pixabay,usgs,noaa,"
                         "commons,nasa,archive,youtube, "
                         "youtubekids (Opt-in-Weg: einbetten, nur made-for-kids), "
                         "ytkatalog (EMPFOHLEN fuer YouTube: lokaler Kanalkatalog, "
                         "verbraucht kein Suchkontingent), ytchannels (Live-Suche je Kanal, "
                         "max. 100 Suchanfragen/Tag) — die drei nicht im Default")
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    ap.add_argument("--xlsx", default=os.path.join(desktop, "video_kandidaten.xlsx"),
                    help="Ziel-Excel (Standardausgabe, default Desktop)")
    ap.add_argument("--no-thumbs", action="store_true",
                    help="Keine Vorschaubilder ins Excel einbetten (schneller)")
    ap.add_argument("--no-probe", action="store_true",
                    help="Kein ffprobe (schneller, aber ohne echte Laenge/Sprach-Erkennung)")
    ap.add_argument("--html", help="Optional: zusaetzlich HTML-Report ablegen")
    ap.add_argument("--json", help="Optional: Kandidaten zusaetzlich als JSON ablegen")
    args = ap.parse_args()

    if args.topics_file:
        with open(args.topics_file, encoding="utf-8") as f:
            raw = [ln.strip() for ln in f if ln.strip() and not ln.startswith("#")]
    else:
        raw = [t.strip() for t in args.topics.split(",") if t.strip()]
    # Alias-Syntax "Anzeigename::Suchbegriff" (fuer kurze/mehrdeutige dt. Woerter)
    topics = []  # Liste (anzeige, suchbegriff)
    for entry in raw:
        if "::" in entry:
            disp, query = entry.split("::", 1)
            topics.append((disp.strip(), query.strip()))
        else:
            topics.append((entry, entry))

    srcs = {s.strip() for s in args.sources.split(",")}
    yt_key = os.environ.get("YT_API_KEY")
    pexels_key = os.environ.get("PEXELS_API_KEY")
    pixabay_key = os.environ.get("PIXABAY_API_KEY")
    if ("youtube" in srcs or "youtubekids" in srcs) and not yt_key:
        print("[i] YouTube uebersprungen (keine YT_API_KEY gesetzt -> Google Cloud: "
              "YouTube Data API v3 aktivieren, Key in .env als YT_API_KEY).")
    if "pexels" in srcs and not pexels_key:
        print("[i] Pexels uebersprungen (keine PEXELS_API_KEY gesetzt) -> gratis: pexels.com/api")
    if "pixabay" in srcs and not pixabay_key:
        print("[i] Pixabay uebersprungen (keine PIXABAY_API_KEY gesetzt) -> gratis: pixabay.com/api/docs")
    if not args.no_probe and not FFPROBE:
        print("[i] ffprobe nicht gefunden -> ohne echte Laenge/Sprach-Erkennung.")

    ch = (args.channels or "").strip().lower()
    if ch == "all":
        allow_channels = None
        print("[i] YouTube-Kanalfilter AUS -> Werbefreiheit NICHT garantiert.")
    elif ch == "oer":
        allow_channels = OER_CHANNELS
        print("[i] YouTube nur oeffentlich-rechtliche Kanaele (gesetzlich werbefrei).")
    else:
        allow_channels = [x.strip() for x in args.channels.split(",") if x.strip()]
    # Kanal-Allowlist ist selbst das Vertrauenssignal -> Kids-Flag dann nicht zwingend
    require_kids = allow_channels is None

    yt_channels = []
    if ("ytchannels" in srcs or "ytkatalog" in srcs) and yt_key:
        tiers = ("oer",) if args.tier == "oer" else ("oer", "komm")
        wanted = [(n, t) for n, t in YT_CHANNEL_REGISTRY if t in tiers]
        print(f"[i] Kanal-Quelle: {len(wanted)} Kanaele (Tier: {args.tier}) — "
              f"loese IDs auf (gecacht)...")
        ids = resolve_channel_ids([n for n, _ in wanted], yt_key)
        yt_channels = [(n, t, ids.get(n, "")) for n, t in wanted]
        ok = sum(1 for _, _, c in yt_channels if c)
        print(f"[i] {ok}/{len(yt_channels)} Kanal-IDs verfuegbar.")
        if args.tier == "beide":
            print("[!] Tier 'beide': kommerzielle Kanaele haben WERBUNG (nicht abschaltbar).")
    if not require_kids:
        print("[i] Kanal-Allowlist aktiv -> made-for-kids-Flag nicht zwingend "
              "(Kanal = Vertrauenssignal).")

    results = {}
    for disp, query in topics:
        print(f"[+] Suche: {disp}  (Begriff: {query})")
        cands = []
        probe = not args.no_probe
        if "terrax" in srcs:
            r, err = search_terrax(query, args.max, probe=probe)
            cands += r
            if err: print("    " + err)
        if "usgs" in srcs:
            r, err = search_usgs(query, args.max, probe=probe)
            cands += r
            if err: print("    " + err)
        if "noaa" in srcs:
            r, err = search_noaa(query, args.max, probe=probe)
            cands += r
            if err: print("    " + err)
        if "gailhampshire" in srcs:
            r, err = search_gailhampshire(query, args.max, probe=probe)
            cands += r
            if err: print("    " + err)
        if "forstmeier" in srcs:
            r, err = search_forstmeier(query, args.max, probe=probe)
            cands += r
            if err: print("    " + err)
        if "esa" in srcs:
            r, err = search_esa(query, args.max, probe=probe)
            cands += r
            if err: print("    " + err)
        if "pexels" in srcs and pexels_key:
            r, err = search_pexels(query, args.max, pexels_key)
            cands += r
            if err: print("    " + err)
        if "pixabay" in srcs and pixabay_key:
            r, err = search_pixabay(query, args.max, pixabay_key)
            cands += r
            if err: print("    " + err)
        if "commons" in srcs:
            r, err = search_commons(query, args.max, probe=probe)
            cands += r
            if err: print("    " + err)
        if "nasa" in srcs:
            r, err = search_nasa(query, args.max, probe=probe)
            cands += r
            if err: print("    " + err)
        if "archive" in srcs:
            r, err = search_archive(query, args.max, probe=probe)
            cands += r
            if err: print("    " + err)
        if "youtube" in srcs and yt_key:
            r, err = search_youtube_cc(query, args.max, args.lang, yt_key)
            cands += r
            if err: print("    " + err)
        if "ytkatalog" in srcs and yt_channels:   # offline im lokalen Kanalkatalog
            r, err = search_catalog(query, yt_channels, yt_key,
                                    per_channel=args.per_channel,
                                    refresh=args.katalog_refresh,
                                    katalog_max=args.katalog_max)
            cands += r
            if err: print("    " + err)
        if "ytchannels" in srcs and yt_channels:  # gezielte Suche IN vertrauten Kanaelen
            r, err = search_youtube_channels(query, yt_key, yt_channels,
                                             per_channel=args.per_channel, lang=args.lang,
                                             strict=not args.locker)
            cands += r
            if err: print("    " + err)
        if "youtubekids" in srcs and yt_key:  # Opt-in-Weg: einbetten, nur made-for-kids
            r, err = search_youtube_kids(query, args.max, args.lang, yt_key,
                                         allow_channels=allow_channels,
                                         require_kids=require_kids)
            cands += r
            if err: print("    " + err)
        # Laengenfilter (unbekannte Laenge bleibt drin)
        if args.min_sec or args.max_sec:
            def _in_range(c):
                s = c.get("dur_sec")
                if s is None:
                    return True
                if args.min_sec and s < args.min_sec:
                    return False
                if args.max_sec and s > args.max_sec:
                    return False
                return True
            before = len(cands)
            cands = [c for c in cands if _in_range(c)]
            print(f"    Laengenfilter {args.min_sec}-{args.max_sec}s: {before} -> {len(cands)}")
        # Bildanalyse nur wenn nach 'warm' sortiert wird (spart Downloads)
        if args.sort == "warm":
            for c in cands:
                if c.get("thumb"):
                    wp, dp = analyze_thumb(c["thumb"])
                    c["warm_pct"], c["dark_pct"] = wp, dp
        # Sortierung: 1) Lizenz (gruen) 2) Sprach-Eignung 3) --sort-Signal
        _order = {"green": 0, "embed": 1, "yellow": 1, "red": 2}
        _third = {
            "relevanz": lambda c: 0,  # Quellenreihenfolge beibehalten (stabil)
            "aufloesung": lambda c: -c.get("res_px", 0),
            "beliebtheit": lambda c: -c.get("popularity", 0),
            "warm": lambda c: -(c.get("warm_pct") or 0),
        }[args.sort]
        cands.sort(key=lambda c: (_order.get(c["flag"], 1), c.get("lang_rank", 1), _third(c)))
        # Top-N-Vorauswahl je Thema (nach Sortierung = beste zuerst)
        if args.top and len(cands) > args.top:
            cands = cands[:args.top]
        results[disp] = cands
        print(f"    {len(cands)} Kandidaten")

    xlsx_path = args.xlsx
    try:
        build_xlsx(results, xlsx_path, embed_thumbs=not args.no_thumbs)
    except PermissionError:
        base, ext = os.path.splitext(args.xlsx)
        xlsx_path = base + "_neu" + ext
        print(f"[i] {args.xlsx} ist gesperrt (in Excel offen?) -> speichere als {xlsx_path}")
        build_xlsx(results, xlsx_path, embed_thumbs=not args.no_thumbs)
    print(f"\n[OK] Excel (klickbar + kommentierbar): {xlsx_path}")
    if args.html:
        build_html(results, args.html)
        print(f"[OK] HTML: {args.html}")
    if args.json:
        with open(args.json, "w", encoding="utf-8") as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        print(f"[OK] JSON: {args.json}")


if __name__ == "__main__":
    main()
