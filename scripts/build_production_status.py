#!/usr/bin/env python3
"""
build_production_status.py — Haengt ein Sheet "Produktion" an catalog_review_master.xlsx.

Reiner LESE-Scan von articles/batch_output/ (aendert nichts am Datenfluss):
  generiert   = articles/{slug}_l{N}.json
  lektoriert  = lektorat/lektorat_{slug}_l{N}.json
  vertont     = audio/{slug}_l{N}_artikel.wav
  (ausgeliefert = spaeter, wenn ein Deployment-Manifest existiert)

slug = thema.lower().replace(" ","_").replace("/","_")   (identisch zu run_batch.py)

Zeigt je Thema/Stufe den hoechsten erreichten Meilenstein, farbcodiert. Gelistet werden
nur Themen mit mind. einem Artefakt (der aktive Produktions-Stand); die Titelzeile nennt
die Gesamtzahl der Include-Themen als Nenner.

Muss NACH build_master.py laufen (build_master schreibt die Mappe frisch und wuerde ein
vorher angehaengtes Produktion-Sheet sonst verwerfen). Master darf nicht in Excel offen sein.

Aufruf: python -X utf8 scripts/build_production_status.py
"""

import pathlib
import sys
from datetime import datetime, timezone

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

REPO   = pathlib.Path(__file__).parent.parent
MASTER = REPO / "catalog_review_master.xlsx"
BATCH  = REPO / "articles" / "batch_output"
ART    = BATCH / "articles"
LEKT   = BATCH / "lektorat"
AUDIO  = BATCH / "audio"

STUFEN = (1, 2, 3)

# Meilenstein-Stufen (hoeher = weiter)
M_NONE, M_GEN, M_LEKT, M_TON = 0, 1, 2, 3
M_LABEL = {M_NONE: "", M_GEN: "generiert", M_LEKT: "lektoriert", M_TON: "vertont"}
M_FILL = {
    M_NONE: None,
    M_GEN:  PatternFill("solid", fgColor="FFF2CC"),   # gelb
    M_LEKT: PatternFill("solid", fgColor="DDEBF7"),   # blau
    M_TON:  PatternFill("solid", fgColor="C6EFCE"),   # gruen
}
FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FONT_HEADER = Font(color="FFFFFF", bold=True)
FONT_BOLD   = Font(bold=True)
FILL_TITLE  = PatternFill("solid", fgColor="D9D9D9")


def slug(thema: str) -> str:
    return thema.lower().replace(" ", "_").replace("/", "_")


def milestone(sl: str, n: int) -> int:
    """Hoechster erreichter Meilenstein fuer Thema-Slug + Stufe."""
    gen = (ART  / f"{sl}_l{n}.json").exists()
    lek = (LEKT / f"lektorat_{sl}_l{n}.json").exists()
    ton = (AUDIO / f"{sl}_l{n}_artikel.wav").exists()
    if ton:
        return M_TON
    if lek:
        return M_LEKT
    if gen:
        return M_GEN
    return M_NONE


def read_master_topics() -> list[dict]:
    """thema/themengebiet/eignung aus dem Review-Sheet."""
    wb = openpyxl.load_workbook(MASTER, read_only=True, data_only=True)
    ws = wb["Review"]
    rows = ws.iter_rows(values_only=True)
    hdr = list(next(rows))
    idx = {h: i for i, h in enumerate(hdr) if h}
    out = []
    for r in rows:
        thema = r[idx["thema"]] if "thema" in idx else None
        if not thema:
            continue
        out.append({
            "thema":        str(thema).strip(),
            "themengebiet": str(r[idx["themengebiet"]] or "") if "themengebiet" in idx else "",
            "eignung":      str(r[idx["eignung"]] or "") if "eignung" in idx else "",
        })
    wb.close()
    return out


def main() -> None:
    if not MASTER.exists():
        sys.exit(f"FEHLT: {MASTER}")
    lock = MASTER.parent / f"~${MASTER.name}"
    if lock.exists():
        sys.exit(f"Master ist in Excel geoeffnet ({lock.name}) — bitte schliessen und erneut ausfuehren.")

    topics = read_master_topics()
    n_include = sum(1 for t in topics if t["eignung"] == "include")

    # Status je Thema
    active = []   # Themen mit mind. einem Artefakt
    for t in topics:
        sl = slug(t["thema"])
        levels = [milestone(sl, n) for n in STUFEN]
        if any(lv > M_NONE for lv in levels):
            t["levels"] = levels
            active.append(t)

    active.sort(key=lambda t: (t["themengebiet"] or "zzz", t["thema"]))

    # Zaehlungen fuer die Titelzeile
    def count_at_least(stufe_idx: int, lvl: int) -> int:
        return sum(1 for t in active if t["levels"][stufe_idx] >= lvl)
    n_gen  = sum(1 for t in active if any(lv >= M_GEN  for lv in t["levels"]))
    n_lekt = sum(1 for t in active if any(lv >= M_LEKT for lv in t["levels"]))
    n_ton  = sum(1 for t in active if any(lv >= M_TON  for lv in t["levels"]))
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    # Master oeffnen, Produktion-Sheet neu schreiben
    wb = openpyxl.load_workbook(MASTER)
    if "Produktion" in wb.sheetnames:
        del wb["Produktion"]
    ws = wb.create_sheet("Produktion")

    COLS = ["thema", "themengebiet", "eignung", "S1", "S2", "S3"]

    # Titelzeile (Row 1)
    title = (f"Produktion — Stand {stamp} · {len(active)} von {n_include} Include-Themen aktiv "
             f"· generiert {n_gen} · lektoriert {n_lekt} · vertont {n_ton} "
             f"· Legende: gelb=generiert, blau=lektoriert, gruen=vertont")
    ws.cell(row=1, column=1, value=title).font = FONT_BOLD
    ws.cell(row=1, column=1).fill = FILL_TITLE
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))

    # Header (Row 2)
    for ci, col in enumerate(COLS, 1):
        c = ws.cell(row=2, column=ci, value=col)
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = Alignment(horizontal="center")

    # Daten ab Row 3
    for ri, t in enumerate(active, 3):
        ws.cell(row=ri, column=1, value=t["thema"])
        ws.cell(row=ri, column=2, value=t["themengebiet"])
        ws.cell(row=ri, column=3, value=t["eignung"])
        for si, lv in enumerate(t["levels"]):
            c = ws.cell(row=ri, column=4 + si, value=M_LABEL[lv])
            if M_FILL[lv]:
                c.fill = M_FILL[lv]
            c.alignment = Alignment(horizontal="center")

    # Layout
    widths = {"A": 30, "B": 30, "C": 10, "D": 12, "E": 12, "F": 12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A3"
    last = max(2, len(active) + 2)
    ws.auto_filter.ref = f"A2:F{last}"

    wb.save(MASTER)
    print(f"Sheet 'Produktion' geschrieben: {len(active)} aktive Themen "
          f"(von {n_include} Includes) — generiert {n_gen} / lektoriert {n_lekt} / vertont {n_ton}.")
    for t in active:
        lv = "/".join(M_LABEL[x] or "–" for x in t["levels"])
        print(f"  {t['thema']:24.24} S1/S2/S3: {lv}")


if __name__ == "__main__":
    main()
