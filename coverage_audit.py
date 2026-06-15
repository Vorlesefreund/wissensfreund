#!/usr/bin/env python3
"""
Coverage Audit: Klexikon + Pflichtliste + LLM-Audit pro Gebiet
→ coverage_gaps_*.json + catalog_review_audit.xlsx
"""

import json, re, difflib, os, sys, time
from collections import defaultdict, Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
load_dotenv()
import anthropic

# ── Dateipfade ───────────────────────────────────────────────────────────────
CATALOG_FULL    = "catalog_full.json"
CATALOG_RESERVE = "catalog_reserve.json"
KLEXIKON_FILE   = "klexikon_appeal_quartil.json"
MASTER_XLSX     = "catalog_review_master.xlsx"
OUT_XLSX        = "catalog_review_audit.xlsx"
OUT_KLEXIKON    = "coverage_gaps_klexikon.json"
OUT_PFLICHT     = "coverage_gaps_pflichtliste.json"
OUT_LLM         = "coverage_gaps_llm.json"

MODEL           = "claude-haiku-4-5-20251001"

FILL_GREEN  = PatternFill("solid", fgColor="D6F0D6")
FILL_ORANGE = PatternFill("solid", fgColor="FFD580")

GEBIETE = [
    "Tiere",
    "Pflanzen & Pilze",
    "Geschichte & Epochen",
    "Gesellschaft, Berufe & Zusammenleben",
    "Grundbegriffe",
    "Kunst, Musik & Literatur",
    "Länder & Kontinente",
    "Menschlicher Körper & Gesundheit",
    "Märchen, Mythologie & Fabelwesen",
    "Naturräume & Landschaften",
    "Naturwissenschaft & Biologie-Konzepte",
    "Religion, Feste & Bräuche",
    "Sport & Spiele",
    "Technik, Maschinen & Fahrzeuge",
    "Erde, Wetter & Naturphänomene",
    "Weltall & Astronomie",
    "Weltstädte & Wahrzeichen",
    "Deutsche Städte",
    "Berühmte Personen",
    "Essen & Alltag",
]

# ── Normalisierung ───────────────────────────────────────────────────────────
def norm(s: str) -> str:
    s = s.lower().strip()
    s = re.sub(r"\s*\(.*?\)", "", s).strip()
    s = re.sub(r"\s+", " ", s)
    return s

def norm_variants(s: str) -> set:
    """Erzeugt normalisierte Varianten inkl. häufiger dt. Plural-Endungen."""
    n = norm(s)
    vs = {n}
    for sfx in ("nen", "inen", "innen", "en", "n", "e", "er", "s"):
        if n.endswith(sfx) and len(n) > len(sfx) + 2:
            vs.add(n[: -len(sfx)])
    return vs

# ── Katalog laden ────────────────────────────────────────────────────────────
print("Lade Katalog …")
catalog_full    = json.load(open(CATALOG_FULL,    encoding="utf-8"))
catalog_reserve = json.load(open(CATALOG_RESERVE, encoding="utf-8"))
all_entries     = catalog_full + catalog_reserve

catalog_themen    = [x["thema"] for x in all_entries]
catalog_norm_list = [norm(t) for t in catalog_themen]
catalog_norm_set  = set(catalog_norm_list)

# Gebiet-Index
gebiet_themen: dict[str, list[str]] = defaultdict(list)
for x in all_entries:
    g = x.get("themengebiet", "")
    if g:
        gebiet_themen[g].append(x["thema"])

def in_catalog(thema: str) -> bool:
    return bool(norm_variants(thema) & catalog_norm_set)

def fuzzy_matches(thema: str, n: int = 3, cutoff: float = 0.70) -> list[tuple[str, float]]:
    """Gibt [(original_thema, ratio)] zurück."""
    nv = norm(thema)
    close = difflib.get_close_matches(nv, catalog_norm_list, n=n, cutoff=cutoff)
    result = []
    for c in close:
        ratio = difflib.SequenceMatcher(None, nv, c).ratio()
        orig  = catalog_themen[catalog_norm_list.index(c)]
        result.append((orig, ratio))
    return result

# ── Heuristik: Gebiet schätzen (für Klexikon/Pflicht-Items ohne Gebiet) ──────
_KMAP = {
    "Tiere": [
        "tier", "hund", "katze", "vogel", "fisch", "pferd", "dino", "affe",
        "löwe", "elefant", "bär", "wolf", "fuchs", "hai", "wal", "schlange",
        "frosch", "spinne", "biene", "schmetterling", "adler", "pinguin",
        "krokodil", "gorilla", "zebra", "giraffe", "nashorn", "nilpferd",
        "leopard", "tiger", "gepard", "luchs", "eule", "rabe", "möwe",
        "otter", "biber", "igel", "hamster", "maus", "ratte", "kaninchen",
        "hase", "eichhörnchen", "hirsch", "reh", "wildschwein", "ziege",
        "schaf", "kuh", "schwein", "hahn", "huhn", "ente", "gans", "taube",
        "lama", "alpaka", "flamingo", "koala", "känguru", "fledermaus",
        "ameise", "käfer", "libelle", "tausendfüssler", "muschel", "krebs",
        "qualle", "oktopus", "tintenfisch", "delfin", "seepferdchen", "schildkröte",
    ],
    "Pflanzen & Pilze": [
        "pflanze", "baum", "blume", "pilz", "gras", "rose", "tulpe", "eiche",
        "birke", "fichte", "kiefer", "tanne", "buche", "kastanie", "linde",
        "walnuss", "bambus", "kaktus", "alge", "moos", "farn", "efeu",
        "sonnenblume", "getreide", "weizen", "mais", "kartoffel",
    ],
    "Geschichte & Epochen": [
        "römer", "röm", "mittelalter", "ritter", "pharao", "wikinger",
        "indianer", "revolution", "kaiser", "weltkrieg", "napoleon",
        "cäsar", "pyramide", "antike", "griech", "ägypt", "babylon",
        "perser", "kreuzzug", "reform", "renaissance", "aufklärung",
        "kolonial", "steinzeit", "bronzezeit", "eisenzeit", "eiszeit",
        "prähistor", "urmensch", "altertum", "altes griechenland",
        "angelsachsen", "kelten", "germanen", "azteken", "maya", "inka",
        "christoph kolumbus", "entdeckung", "dreißigjähr",
    ],
    "Gesellschaft, Berufe & Zusammenleben": [
        "familie", "freundschaft", "schule", "kindergarten", "geburtstag",
        "beruf", "arzt", "feuerwehr", "polizei", "lehrer", "bauer",
        "demokratie", "wahl", "partei", "recht", "gesetz", "parlament",
        "gemeinde", "bundesland", "armut", "flüchtling", "migration",
        "krankheit", "behinderung",
    ],
    "Grundbegriffe": [
        "farbe", "zahl", "form", "alphabet", "lesen", "schreib", "rechnen",
        "uhr", "kalender", "woche", "monat", "jahreszeit", "frühling",
        "sommer", "herbst", "winter", "adjektiv", "verb", "substantiv",
        "adverb", "grammatik", "sprache", "wort",
    ],
    "Kunst, Musik & Literatur": [
        "musik", "kunst", "literatur", "maler", "musiker", "buch", "roman",
        "gedicht", "theater", "oper", "tanz", "ballet", "film", "kino",
        "foto", "architektur", "skulptur", "gemälde", "beethoven", "mozart",
        "bach", "shakespeare", "goethe", "schiller",
    ],
    "Länder & Kontinente": [
        "land", "kontinent", "europa", "asien", "afrika", "australien",
        "nordamerika", "südamerika", "antarktis", "arabien", "arabisch",
        "china", "indien", "russland", "usa", "brasilien", "kanada",
        "mexiko", "türkei", "japan", "korea", "vietnam", "thailand",
        "ägypten", "nigeria", "kenia", "marokko",
    ],
    "Menschlicher Körper & Gesundheit": [
        "körper", "gehirn", "herz", "lunge", "niere", "leber", "magen",
        "darm", "blut", "knochen", "muskel", "haut", "zähne", "auge",
        "ohr", "nase", "zelle", "impfung", "allergie", "krebs",
        "vitaminen", "ernährung", "pubertät",
    ],
    "Märchen, Mythologie & Fabelwesen": [
        "märchen", "mythos", "drachen", "einhorn", "meerjungfrau",
        "zwerg", "elfe", "fee", "schneewittchen", "rotkäppchen",
        "aschenputtel", "hänsel", "dornröschen", "rapunzel",
        "rumpelstilz", "bremer", "froschkönig", "griechi", "nordisch",
        "zeus", "odin", "thor", "herkules", "odyssee",
    ],
    "Naturräume & Landschaften": [
        "wüste", "dschungel", "regenwald", "savanne", "steppe", "tundra",
        "taiga", "prairie", "atlantik", "pazifik", "arktis", "antarktis",
        "nordsee", "ostsee", "alpen", "schwarzwald", "sahara",
        "amazonas", "nil", "ural",
    ],
    "Naturwissenschaft & Biologie-Konzepte": [
        "atom", "molekül", "chemie", "physik", "energie", "elektrizität",
        "magnet", "licht", "optik", "wärme", "kälte", "aggregat",
        "evolution", "ökosystem", "fotosynthese", "nahrungskette",
        "zelle", "gen", "dna", "klon", "astronomie", "erdkunde",
        "geologie",
    ],
    "Religion, Feste & Bräuche": [
        "religion", "christentum", "islam", "judentum", "buddhismus",
        "hinduismus", "weihnachten", "ostern", "karneval", "silvester",
        "ramadan", "chanukka", "tempel", "kirche", "moschee", "synagoge",
        "abendmahl", "taufe", "bibel", "koran", "torah",
    ],
    "Sport & Spiele": [
        "sport", "fußball", "basketball", "tennis", "schwimm", "leichtathletik",
        "turnen", "handball", "volleyball", "hockey", "golf", "boxen",
        "ringen", "judo", "karate", "olympia", "weltmeister", "spiel",
        "schach", "lacrosse",
    ],
    "Technik, Maschinen & Fahrzeuge": [
        "auto", "flugzeug", "zug", "schiff", "fahrrad", "motorrad",
        "rakete", "u-boot", "hubschrauber", "roboter", "computer",
        "telefon", "smartphone", "fernseher", "radio", "kamera",
        "elektromotor", "dampfmaschine", "druckmaschine", "app",
    ],
    "Erde, Wetter & Naturphänomene": [
        "wetter", "regen", "schnee", "eis", "sturm", "gewitter",
        "regenbogen", "tornado", "taifun", "erdbeben", "tsunami",
        "vulkan", "meteor", "komet", "klima", "atmosphäre",
        "abfall", "recycling", "umwelt",
    ],
    "Weltall & Astronomie": [
        "weltall", "stern", "planet", "mond", "sonne", "galaxie",
        "milchstraße", "schwarzes loch", "komet", "asteroid", "meteorit",
        "teleskop", "apollo", "astronaut", "iss",
    ],
    "Weltstädte & Wahrzeichen": [
        "paris", "london", "new york", "tokio", "dubai", "sydney",
        "rom", "athen", "kairo", "moskau", "peking", "mumbai",
        "wahrzeichen", "eiffelturm", "big ben", "freiheitsstatue",
        "kolosseum",
    ],
    "Deutsche Städte": [
        "münchen", "hamburg", "köln", "frankfurt", "stuttgart", "düsseldorf",
        "dortmund", "essen", "bremen", "hannover", "nürnberg", "leipzig",
        "dresden", "bonn", "mainz", "kiel", "erfurt",
    ],
    "Berühmte Personen": [
        "könig", "queen", "präsident", "papst", "einstein", "newton",
        "darwin", "kopernikus", "galilei", "freud", "picasso", "da vinci",
        "michelangelo", "rembrandt", "beethoven", "mozart", "bach",
        "shakespeare", "dickens", "hugo", "tolstoi",
    ],
    "Essen & Alltag": [
        "essen", "kochen", "brot", "butter", "milch", "käse", "eier",
        "obst", "gemüse", "kuchen", "schokolade", "speiseeis", "pizza",
        "pasta", "suppe", "fleisch", "fisch", "reis", "einkauf",
    ],
}

def guess_gebiet(thema: str) -> str:
    t = norm(thema)
    for gebiet, keywords in _KMAP.items():
        for kw in keywords:
            if kw in t:
                return gebiet
    return ""

# ── Teil 1: Klexikon-Abgleich ─────────────────────────────────────────────────
print("\n[Teil 1] Klexikon-Abgleich …")
klexikon = json.load(open(KLEXIKON_FILE, encoding="utf-8"))
klexikon.sort(key=lambda x: (x["quartil"], x["klexikon_titel"]))

gaps_klexikon: list[dict] = []
for item in klexikon:
    titel = item["klexikon_titel"]
    if not in_catalog(titel):
        gaps_klexikon.append({
            "thema":       titel,
            "themengebiet": guess_gebiet(titel),
            "quelle":      "klexikon",
            "detail":      f"Q{item['quartil']}",
        })

print(f"  → {len(klexikon)} Klexikon-Artikel, {len(gaps_klexikon)} Lücken")
json.dump(gaps_klexikon, open(OUT_KLEXIKON, "w", encoding="utf-8"),
          ensure_ascii=False, indent=2)

# ── Teil 2: Pflichtliste ─────────────────────────────────────────────────────
print("\n[Teil 2] Pflichtliste …")
PFLICHT: list[tuple[str, str]] = [
    # (thema, themengebiet)
    # Tiere – bekannte Grundbegriffe
    ("Indianer", "Geschichte & Epochen"),
    ("Flamingo", "Tiere"),
    ("Koala", "Tiere"),
    ("Känguru", "Tiere"),
    ("Lama", "Tiere"),
    ("Alpaka", "Tiere"),
    ("Fledermaus", "Tiere"),
    ("Igel", "Tiere"),
    ("Eichhörnchen", "Tiere"),
    ("Hase", "Tiere"),
    ("Kaninchen", "Tiere"),
    ("Hamster", "Tiere"),
    ("Meerschweinchen", "Tiere"),
    ("Wellensittich", "Tiere"),
    ("Goldfisch", "Tiere"),
    ("Kuh", "Tiere"),
    ("Schwein", "Tiere"),
    ("Schaf", "Tiere"),
    ("Ziege", "Tiere"),
    ("Esel", "Tiere"),
    ("Hahn", "Tiere"),
    # Gesellschaft & Alltag
    ("Familie", "Gesellschaft, Berufe & Zusammenleben"),
    ("Freundschaft", "Gesellschaft, Berufe & Zusammenleben"),
    ("Schule", "Gesellschaft, Berufe & Zusammenleben"),
    ("Kindergarten", "Gesellschaft, Berufe & Zusammenleben"),
    ("Geburtstag", "Gesellschaft, Berufe & Zusammenleben"),
    ("Urlaub", "Gesellschaft, Berufe & Zusammenleben"),
    ("Taschengeld", "Gesellschaft, Berufe & Zusammenleben"),
    # Jahreszeiten / Wetter
    ("Jahreszeiten", "Grundbegriffe"),
    ("Frühling", "Grundbegriffe"),
    ("Sommer", "Grundbegriffe"),
    ("Herbst", "Grundbegriffe"),
    ("Winter", "Grundbegriffe"),
    ("Regen", "Erde, Wetter & Naturphänomene"),
    ("Schnee", "Erde, Wetter & Naturphänomene"),
    ("Gewitter", "Erde, Wetter & Naturphänomene"),
    ("Regenbogen", "Erde, Wetter & Naturphänomene"),
    # Grundbegriffe
    ("Farben", "Grundbegriffe"),
    ("Zahlen", "Grundbegriffe"),
    ("Alphabet", "Grundbegriffe"),
    ("Lesen", "Grundbegriffe"),
    ("Schreiben", "Grundbegriffe"),
    ("Rechnen", "Grundbegriffe"),
    ("Uhr", "Grundbegriffe"),
    ("Kalender", "Grundbegriffe"),
    # Essen
    ("Brot", "Essen & Alltag"),
    ("Milch", "Essen & Alltag"),
    ("Käse", "Essen & Alltag"),
    ("Eier", "Essen & Alltag"),
    ("Obst", "Essen & Alltag"),
    ("Gemüse", "Essen & Alltag"),
    ("Kuchen", "Essen & Alltag"),
    ("Schokolade", "Essen & Alltag"),
    ("Speiseeis", "Essen & Alltag"),
    ("Salat", "Essen & Alltag"),
    ("Suppe", "Essen & Alltag"),
    ("Reis", "Essen & Alltag"),
    # Körper
    ("Zähne", "Menschlicher Körper & Gesundheit"),
    ("Haare", "Menschlicher Körper & Gesundheit"),
    ("Augen", "Menschlicher Körper & Gesundheit"),
    ("Ohren", "Menschlicher Körper & Gesundheit"),
    ("Schlaf", "Menschlicher Körper & Gesundheit"),
    ("Traum", "Menschlicher Körper & Gesundheit"),
    # Technik
    ("Fahrrad", "Technik, Maschinen & Fahrzeuge"),
    ("Motorrad", "Technik, Maschinen & Fahrzeuge"),
    ("U-Boot", "Technik, Maschinen & Fahrzeuge"),
    ("Smartphone", "Technik, Maschinen & Fahrzeuge"),
    ("Fernseher", "Technik, Maschinen & Fahrzeuge"),
    ("Radio", "Technik, Maschinen & Fahrzeuge"),
    ("Kamera", "Technik, Maschinen & Fahrzeuge"),
    ("Roboter", "Technik, Maschinen & Fahrzeuge"),
    # Natur-Konzepte
    ("Fotosynthese", "Naturwissenschaft & Biologie-Konzepte"),
    ("Nahrungskette", "Naturwissenschaft & Biologie-Konzepte"),
    ("Ökosystem", "Naturwissenschaft & Biologie-Konzepte"),
    ("Artenvielfalt", "Naturwissenschaft & Biologie-Konzepte"),
    ("Recycling", "Naturwissenschaft & Biologie-Konzepte"),
    ("Klimawandel", "Naturwissenschaft & Biologie-Konzepte"),
    # Geschichte
    ("Ritter", "Geschichte & Epochen"),
    ("Mittelalter", "Geschichte & Epochen"),
    ("Wikinger", "Geschichte & Epochen"),
    ("Pharao", "Geschichte & Epochen"),
    ("Mumie", "Geschichte & Epochen"),
    ("Pyramide", "Geschichte & Epochen"),
    ("Römer", "Geschichte & Epochen"),
    # Geographie / Natur
    ("Atlantik", "Naturräume & Landschaften"),
    ("Pazifik", "Naturräume & Landschaften"),
    ("Arktis", "Naturräume & Landschaften"),
    ("Antarktis", "Naturräume & Landschaften"),
    ("Dschungel", "Naturräume & Landschaften"),
    ("Regenwald", "Naturräume & Landschaften"),
    ("Savanne", "Naturräume & Landschaften"),
    ("Steppe", "Naturräume & Landschaften"),
    ("Nordsee", "Naturräume & Landschaften"),
    ("Ostsee", "Naturräume & Landschaften"),
    ("Alpen", "Naturräume & Landschaften"),
    # Weltall
    ("Komet", "Weltall & Astronomie"),
    ("Asteroid", "Weltall & Astronomie"),
    ("Meteorit", "Weltall & Astronomie"),
    ("Milchstraße", "Weltall & Astronomie"),
    # Märchen & Mythen
    ("Drachen", "Märchen, Mythologie & Fabelwesen"),
    ("Einhorn", "Märchen, Mythologie & Fabelwesen"),
    ("Meerjungfrau", "Märchen, Mythologie & Fabelwesen"),
    ("Schneewittchen", "Märchen, Mythologie & Fabelwesen"),
    ("Rotkäppchen", "Märchen, Mythologie & Fabelwesen"),
    ("Aschenputtel", "Märchen, Mythologie & Fabelwesen"),
    ("Hänsel und Gretel", "Märchen, Mythologie & Fabelwesen"),
    ("Dornröschen", "Märchen, Mythologie & Fabelwesen"),
    ("Rapunzel", "Märchen, Mythologie & Fabelwesen"),
    ("Rumpelstilzchen", "Märchen, Mythologie & Fabelwesen"),
    # Religion / Feste
    ("Weihnachten", "Religion, Feste & Bräuche"),
    ("Ostern", "Religion, Feste & Bräuche"),
    ("Silvester", "Religion, Feste & Bräuche"),
    ("Karneval", "Religion, Feste & Bräuche"),
    ("Ramadan", "Religion, Feste & Bräuche"),
]

gaps_pflicht: list[dict] = []
seen_norm = set()
for thema, gebiet in PFLICHT:
    n = norm(thema)
    if n in seen_norm:
        continue
    seen_norm.add(n)
    if not in_catalog(thema):
        gaps_pflicht.append({"thema": thema, "themengebiet": gebiet,
                              "quelle": "pflichtliste", "detail": ""})

print(f"  → {len(PFLICHT)} Kandidaten, {len(gaps_pflicht)} Lücken")
json.dump(gaps_pflicht, open(OUT_PFLICHT, "w", encoding="utf-8"),
          ensure_ascii=False, indent=2)

# ── Teil 3: LLM-Audit pro Gebiet ─────────────────────────────────────────────
print("\n[Teil 3] LLM-Audit (Haiku, parallel) …")
client = anthropic.Anthropic()

SYSTEM = (
    "Du bist Kurator einer deutschen Kinder-Enzyklopädie (Zielgruppe 6–14 Jahre). "
    "Deine Aufgabe: Wichtige Themen benennen, die in einem bestimmten Themengebiet "
    "fehlen. Antworte ausschließlich als JSON-Objekt: {\"fehlende_themen\": [\"...\", ...]}. "
    "Keine Erklärung außerhalb des JSON."
)

def llm_audit_gebiet(gebiet: str) -> list[dict]:
    vorhanden = sorted(gebiet_themen.get(gebiet, []))
    vorhanden_str = ", ".join(vorhanden[:120])  # max ~120 um Token zu sparen
    user_msg = (
        f"Themengebiet: **{gebiet}**\n\n"
        f"Bereits vorhandene Themen ({len(vorhanden)}):\n{vorhanden_str}\n\n"
        "Welche wichtigen Themen fehlen noch? Nenne 12–18 Themen, die für Kinder "
        "6–14 Jahre relevant und spannend sind. Nur Themen, die klar zu diesem "
        "Gebiet gehören und nicht schon in der Liste stehen."
    )
    for attempt in range(3):
        try:
            resp = client.messages.create(
                model=MODEL,
                max_tokens=1024,
                messages=[{"role": "user", "content": user_msg}],
                system=SYSTEM,
            )
            text = resp.content[0].text.strip()
            # JSON extrahieren
            m = re.search(r'\{.*\}', text, re.DOTALL)
            if m:
                data = json.loads(m.group())
                themen = data.get("fehlende_themen", [])
                return [
                    {"thema": t, "themengebiet": gebiet,
                     "quelle": "llm", "detail": ""}
                    for t in themen if isinstance(t, str) and t.strip()
                ]
        except Exception as e:
            print(f"    [!] {gebiet} Versuch {attempt+1}: {e}")
            time.sleep(2 ** attempt)
    return []

gaps_llm_raw: list[dict] = []
with ThreadPoolExecutor(max_workers=5) as pool:
    futures = {pool.submit(llm_audit_gebiet, g): g for g in GEBIETE}
    for fut in as_completed(futures):
        g = futures[fut]
        res = fut.result()
        print(f"  {g}: {len(res)} Vorschläge")
        gaps_llm_raw.extend(res)

# Nur wirklich fehlende behalten
gaps_llm = [x for x in gaps_llm_raw if not in_catalog(x["thema"])]
print(f"  → {len(gaps_llm_raw)} Vorschläge, {len(gaps_llm)} davon wirklich neu")
json.dump(gaps_llm, open(OUT_LLM, "w", encoding="utf-8"), ensure_ascii=False, indent=2)

# ── Alle Kandidaten zusammenführen ────────────────────────────────────────────
print("\nZusammenführen …")
all_gaps: list[dict] = []
seen_combined: set[str] = set()

for gap_list in (gaps_klexikon, gaps_pflicht, gaps_llm):
    for g in gap_list:
        n = norm(g["thema"])
        if n not in seen_combined:
            seen_combined.add(n)
            all_gaps.append(g)

print(f"  Gesamt neue Kandidaten (dedup): {len(all_gaps)}")

# ── Fuzzy-Match für AEHNLICH_VORHANDEN ───────────────────────────────────────
print("Fuzzy-Matching …")
for g in all_gaps:
    matches = fuzzy_matches(g["thema"], n=3, cutoff=0.70)
    g["aehnlich_vorhanden"] = "; ".join(f"{t} ({r:.2f})" for t, r in matches)
    g["mutmassliche_dublette"] = "JA" if any(r >= 0.85 for _, r in matches) else ""

dupl_count = sum(1 for g in all_gaps if g["mutmassliche_dublette"] == "JA")
print(f"  Mutmaßliche Dubletten: {dupl_count}")

# ── Excel bauen ───────────────────────────────────────────────────────────────
print("\nExcel aufbauen …")
src_wb = openpyxl.load_workbook(MASTER_XLSX, data_only=True)
src_ws = src_wb["Review"]

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ── Sheet 1: Review ─────────────────────────────────────────────────────────
ws = wb.create_sheet("Review")

# Quell-Spalten
src_hdr = [c.value for c in next(src_ws.iter_rows(max_row=1))]

# Neue Spalten vorne (nach thema)
NEW_COLS_AFTER_THEMA = ["STATUS_NEU", "QUELLE_AUDIT", "MUTMASSLICHE_DUBLETTE", "AEHNLICH_VORHANDEN"]

# Ziel-Header: thema | STATUS_NEU | QUELLE_AUDIT | MUTT_DUPL | AEHNLICH | rest
thema_idx = src_hdr.index("thema")
cols_before = src_hdr[:thema_idx + 1]                  # inkl. thema
cols_after  = src_hdr[thema_idx + 1:]                  # rest

target_hdr = cols_before + NEW_COLS_AFTER_THEMA + cols_after
ws.append(target_hdr)

# Header-Format
hdr_fill = PatternFill("solid", fgColor="4472C4")
hdr_font = Font(bold=True, color="FFFFFF")
for cell in ws[1]:
    cell.fill = hdr_fill
    cell.font = hdr_font

# AutoFilter + Freeze
ws.auto_filter.ref = f"A1:{get_column_letter(len(target_hdr))}1"
ws.freeze_panes = "B2"

# Neue-Kandidaten-Rows vorbereiten: sortiert nach thema (alphabetisch)
all_gaps_sorted = sorted(all_gaps, key=lambda x: x["thema"].lower())

# Neue Kandidaten einfügen (oben, da alphabetisch gemischt → wir mischen alles)
# Strategie: neue Zeilen + alte Zeilen, BEIDE alphabetisch nach thema sortiert
# ⟹ Erst alle alten Zeilen als dict einlesen, dann zusammen sortieren
old_rows: list[dict] = []
for row in src_ws.iter_rows(min_row=2, values_only=True):
    rd = dict(zip(src_hdr, row))
    if rd.get("thema"):
        old_rows.append(rd)

# Neue Kandidaten in ziel-Dict-Format bringen
def gap_to_row(g: dict) -> dict:
    rd: dict = {h: None for h in target_hdr}
    rd["thema"]               = g["thema"]
    rd["themengebiet"]        = g.get("themengebiet", "")
    rd["STATUS_NEU"]          = "NEU"
    rd["QUELLE_AUDIT"]        = g.get("quelle", "")
    rd["MUTMASSLICHE_DUBLETTE"] = g.get("mutmassliche_dublette", "")
    rd["AEHNLICH_VORHANDEN"]  = g.get("aehnlich_vorhanden", "")
    rd["tier"]                = ""
    return rd

new_rows = [gap_to_row(g) for g in all_gaps_sorted]

# Gemeinsam alphabetisch sortieren
combined = new_rows + [
    {**{h: None for h in target_hdr},
     **rd,
     "STATUS_NEU": "",
     "QUELLE_AUDIT": "",
     "MUTMASSLICHE_DUBLETTE": "",
     "AEHNLICH_VORHANDEN": ""}
    for rd in old_rows
]
combined.sort(key=lambda r: (r.get("thema") or "").lower())

# Schreiben
new_thema_set = {norm(g["thema"]) for g in all_gaps_sorted}
dupl_thema_set = {norm(g["thema"]) for g in all_gaps_sorted
                  if g.get("mutmassliche_dublette") == "JA"}

for rd in combined:
    row_vals = [rd.get(h) for h in target_hdr]
    ws.append(row_vals)
    row_n = ws.max_row
    thema_val = rd.get("thema") or ""
    n_val     = norm(thema_val)
    is_new    = n_val in new_thema_set
    is_dupl   = n_val in dupl_thema_set
    if is_dupl:
        for cell in ws[row_n]:
            cell.fill = FILL_ORANGE
    elif is_new:
        for cell in ws[row_n]:
            cell.fill = FILL_GREEN

# Spaltenbreite
col_widths = {"thema": 35, "themengebiet": 30, "STATUS_NEU": 10,
              "QUELLE_AUDIT": 14, "MUTMASSLICHE_DUBLETTE": 18,
              "AEHNLICH_VORHANDEN": 50, "framing_note": 40,
              "begruendung_eignung": 40, "notiz": 30, "Kommentar": 30}
for i, h in enumerate(target_hdr, 1):
    ws.column_dimensions[get_column_letter(i)].width = col_widths.get(h, 14)

# ── Sheet 2: Audit_Statistik ─────────────────────────────────────────────────
ws2 = wb.create_sheet("Audit_Statistik")
ws2.append(["Kennzahl", "Wert"])
ws2.append(["Gesamt neue Kandidaten", len(all_gaps)])
ws2.append(["Mutmaßliche Dubletten", dupl_count])
ws2.append(["Davon Klexikon-Lücken", sum(1 for g in all_gaps if g["quelle"] == "klexikon")])
ws2.append(["Davon Pflichtliste-Lücken", sum(1 for g in all_gaps if g["quelle"] == "pflichtliste")])
ws2.append(["Davon LLM-Lücken", sum(1 for g in all_gaps if g["quelle"] == "llm")])
ws2.append([])
ws2.append(["Quelle", "Anzahl"])
quelle_cnt = Counter(g["quelle"] for g in all_gaps)
for k, v in sorted(quelle_cnt.items()):
    ws2.append([k, v])
ws2.append([])
ws2.append(["Gebiet", "Anzahl neue Kandidaten"])
gebiet_cnt = Counter(g.get("themengebiet", "Unbekannt") or "Unbekannt"
                     for g in all_gaps)
for k, v in sorted(gebiet_cnt.items(), key=lambda x: -x[1]):
    ws2.append([k, v])

# Header-Format Sheet 2
for cell in ws2[1]:
    cell.fill = hdr_fill
    cell.font = hdr_font
ws2.column_dimensions["A"].width = 42
ws2.column_dimensions["B"].width = 20

# Dubletten-Sheet aus Quelle
ws3 = wb.create_sheet("Dubletten")
src_dup = src_wb["Dubletten"]
dup_hdr = [c.value for c in next(src_dup.iter_rows(max_row=1))]
ws3.append(dup_hdr)
for row in src_dup.iter_rows(min_row=2, values_only=True):
    ws3.append(list(row))
for cell in ws3[1]:
    cell.fill = hdr_fill
    cell.font = hdr_font

wb.save(OUT_XLSX)
print(f"\n✅ Gespeichert: {OUT_XLSX}")

# ── Zusammenfassung ───────────────────────────────────────────────────────────
print("\n═══ ZUSAMMENFASSUNG ═══")
print(f"  Neue Kandidaten gesamt:     {len(all_gaps)}")
print(f"  Davon Klexikon:             {quelle_cnt.get('klexikon',0)}")
print(f"  Davon Pflichtliste:         {quelle_cnt.get('pflichtliste',0)}")
print(f"  Davon LLM:                  {quelle_cnt.get('llm',0)}")
print(f"  Mutmaßliche Dubletten:      {dupl_count}")
print("\n  Top-Gebiete mit Lücken:")
for g, v in sorted(gebiet_cnt.items(), key=lambda x: -x[1])[:8]:
    print(f"    {g:<42} {v}")
